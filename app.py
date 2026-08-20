import os
import io
import re
import json
import time
import uuid
import shutil
import logging
import threading
import subprocess
from collections import Counter
from datetime import datetime
from markupsafe import Markup
from flask import (Flask, render_template, jsonify, request, send_file,
                   redirect, url_for, session, make_response)
from werkzeug.security import generate_password_hash, check_password_hash

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 300 * 1024 * 1024   # 업로드 최대 300MB(약관 ZIP 등)
# 세션 쿠키 서명 키. 배포 시엔 반드시 SECRET_KEY 환경변수로 고정값 지정
# (여러 인스턴스가 같은 키를 써야 로그인 세션이 공유됨).
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-change-me')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')

# ── 접근 제한 (로그인) ─────────────────────────────────────────────────
# 아이디/비밀번호는 환경변수로 주입. 미설정 시 로컬 개발용 기본값(배포 시 반드시 변경).
APP_USER = os.environ.get('APP_USER', 'admin')
APP_PASSWORD = os.environ.get('APP_PASSWORD', 'goun')
if not (os.environ.get('APP_USER') and os.environ.get('APP_PASSWORD')):
    logger.warning('기본 로그인 계정(admin) 사용 중 — 배포 시 APP_USER/APP_PASSWORD 환경변수를 반드시 설정하세요.')

# ── 소속(role)별 계정 ──
#   연금컨설팅팀(=APP_USER) = role 'consulting' : 전체 관리
#   RM(=RM_USER)            = role 'rm'         : 조회 + 다운로드만(업로드·등록·조회실행·기관관리·상품제안관리 불가)
# RM_USER/RM_PASSWORD 미설정 시 RM 로그인 비활성(연금컨설팅팀 계정만 동작).
RM_USER = os.environ.get('RM_USER', '')
RM_PASSWORD = os.environ.get('RM_PASSWORD', '')
# RM 비밀번호 만료 주기(일). 이 기간이 지나면 RM 로그인이 막히고, 연금컨설팅팀이 재설정해야 다시 로그인 가능.
RM_PW_MAX_AGE_DAYS = int(os.environ.get('RM_PW_MAX_AGE_DAYS', '14'))


def _role_for(username: str):
    if username and username == APP_USER:
        return 'consulting'
    if RM_USER and username == RM_USER:
        return 'rm'
    return None


def _env_password_for(username: str):
    if username == APP_USER:
        return APP_PASSWORD
    if RM_USER and username == RM_USER:
        return RM_PASSWORD
    return None

# 화면에서 비밀번호를 변경하면 해시가 여기 저장되고, 이후 로그인은 이 값이 기준이 된다.
# (data/는 배포 시 GCS 버킷 마운트라 재시작·재배포해도 유지됨. 초기화하려면 이 파일 삭제 →
#  다시 APP_PASSWORD 환경변수 값으로 로그인.)
AUTH_FILE = os.path.join(DATA_DIR, 'auth.json')
MIN_PASSWORD_LEN = 8

# 로그인 없이 접근 허용할 엔드포인트
# api_change_password: 로그인 화면에서 비밀번호를 바꿀 수 있게 공개. 대신 아이디+현재 비밀번호를
# 모두 맞게 입력해야만 통과하므로 로그인 자체와 같은 수준의 검증을 거친다.
_PUBLIC_ENDPOINTS = {'login', 'static', 'api_change_password'}

# RM(조회 전용)이 접근할 수 없는 쓰기/실행/관리 엔드포인트.
# (다운로드·조회 GET은 여기 없음 → RM 허용). 서버측 최종 방어선이며 화면 숨김과 별개로 강제된다.
_ADMIN_ONLY_ENDPOINTS = {
    'pension_store_post',        # 원리금 금리 업로드 저장
    'rate_history_append',       # 과거 금리 추이 월간 평균 추가
    'proposal',                  # 상품제안관리 화면
    'proposals_post',            # 이달의 제안상품 등록
    'proposal_meta_post',        # 유니버스/연컨사전확인/판매가능 저장
    'api_notice_post',           # 공지사항 저장(연금컨설팅팀 전용)
    'api_issue_suspension_post', # 발행정지기간 저장(연금컨설팅팀 전용)
    'api_refresh', 'api_refresh_one',   # 신용등급 지금 조회(스크래핑)
    'api_update_rating',         # 신용등급 수정
    'api_acknowledge', 'api_acknowledge_all',   # 변경 확인
    'api_add_institution', 'api_delete_institution',   # 기관 추가/삭제
    'admin_visitors', 'api_visit_stats', 'download_visit_stats',   # 방문자 통계(연금컨설팅팀 전용 관리자)
    'admin_deploy', 'admin_deploy_status',   # 서버 배포/상태(연금컨설팅팀 전용)
    'admin_refresh_bond_rates',  # 시장금리 수동 갱신(연금컨설팅팀 전용)
    'docs_page', 'api_docs_upload', 'api_docs_delete', 'api_docs_bulk',  # 약관·상품설명서 관리(연금컨설팅팀 전용)
}


def _load_auth() -> dict:
    if not os.path.exists(AUTH_FILE):
        return {}
    try:
        with open(AUTH_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        # 파일이 깨져도 로그인 자체는 막지 않는다(환경변수 비밀번호로 대체).
        logger.exception('auth.json 읽기 실패 — APP_PASSWORD로 대체합니다')
        return {}


def _save_auth(data: dict):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(AUTH_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── 방문자 접속 집계 (일자별) ──────────────────────────────────────────
#   논문(AI 도입 후 사용률 ↔ WM게시판 상관관계) 기초자료용.
#   웹/모바일 모든 접속을 집계하되, 두 지표를 함께 저장:
#     - *_total : 페이지 진입 횟수(총 접속수)
#     - *_uids  : 방문자 쿠키(vid) 목록 → len()이 순방문자(하루·기기당 1회, 브라우저 기준 근사)
#   ⚠️ 런타임 데이터 파일 → 배포 시 코드에서 제외(서버 누적 데이터 보호).
VISIT_FILE = os.path.join(DATA_DIR, 'visit_stats.json')
_VISIT_LOCK = threading.Lock()
_MOBILE_UA_RE = re.compile(r'Mobi|Android|iPhone|iPod|Windows Phone|BlackBerry|IEMobile', re.I)


def _load_visits() -> dict:
    if not os.path.exists(VISIT_FILE):
        return {'days': {}}
    try:
        with open(VISIT_FILE, 'r', encoding='utf-8') as f:
            d = json.load(f)
        if not isinstance(d, dict) or 'days' not in d:
            return {'days': {}}
        return d
    except Exception:
        logger.exception('visit_stats.json 읽기 실패')
        return {'days': {}}


def _save_visits(d: dict):
    os.makedirs(DATA_DIR, exist_ok=True)
    tmp = VISIT_FILE + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(d, f, ensure_ascii=False)
    os.replace(tmp, VISIT_FILE)


def _record_visit(resp):
    """메인 페이지(index) 로드 1회당 호출 — 일자별 총 접속수 + 순방문자 집계."""
    try:
        vid = request.cookies.get('vid')
        new_vid = not vid
        if new_vid:
            vid = uuid.uuid4().hex
        dev = 'mob' if _MOBILE_UA_RE.search(request.headers.get('User-Agent', '') or '') else 'web'
        today = datetime.now().strftime('%Y-%m-%d')
        with _VISIT_LOCK:
            data = _load_visits()
            day = data['days'].setdefault(
                today, {'web_total': 0, 'mob_total': 0, 'web_uids': [], 'mob_uids': []})
            day[dev + '_total'] = day.get(dev + '_total', 0) + 1
            uids = day.setdefault(dev + '_uids', [])
            if vid not in uids:
                uids.append(vid)
            _save_visits(data)
        if new_vid:
            # 2년짜리 방문자 식별 쿠키(개인정보 아님: 무작위 id) → 순방문자 추정용
            resp.set_cookie('vid', vid, max_age=60 * 60 * 24 * 730, samesite='Lax')
    except Exception:
        logger.exception('방문자 집계 실패')  # 집계 실패가 페이지 로드를 막지 않도록
    return resp


def _visit_series() -> list:
    """일자 오름차순 집계 시계열. web/mob 총접속·순방문·합계."""
    data = _load_visits()
    out = []
    for day in sorted(data.get('days', {}).keys()):
        d = data['days'][day] or {}
        wt, mt = int(d.get('web_total', 0)), int(d.get('mob_total', 0))
        wu, mu = len(d.get('web_uids', [])), len(d.get('mob_uids', []))
        out.append({'date': day,
                    'web_total': wt, 'mob_total': mt, 'total': wt + mt,
                    'web_uniq': wu, 'mob_uniq': mu, 'uniq': wu + mu})
    return out


def _user_hash(username: str):
    """사용자의 저장된 비밀번호 해시. 신형 auth.json={username:{password_hash}},
    구형={password_hash}(APP_USER 것)도 호환."""
    auth = _load_auth()
    if 'password_hash' in auth and username == APP_USER:   # 구형 포맷
        return auth.get('password_hash')
    u = auth.get(username)
    return u.get('password_hash') if isinstance(u, dict) else None


def verify_password(username: str, pw: str) -> bool:
    """저장된 해시가 있으면 그것으로, 없으면 해당 계정의 환경변수 비밀번호로 검증."""
    stored = _user_hash(username)
    if stored:
        return check_password_hash(stored, pw)
    envpw = _env_password_for(username)
    return envpw is not None and pw == envpw


# ── RM 비밀번호 만료(2주) ─────────────────────────────────────────────
def _rm_pw_set_at():
    """RM 비밀번호가 마지막으로 설정된 시각 문자열(없으면 None)."""
    if not RM_USER:
        return None
    u = _load_auth().get(RM_USER)
    if isinstance(u, dict):
        return u.get('rm_pw_set_at') or u.get('updated')
    return None


def _mark_rm_pw_set_at():
    """RM 비밀번호 설정 시각을 now로 기록(만료 시계 리셋). 비밀번호 해시는 건드리지 않음."""
    if not RM_USER:
        return
    auth = _load_auth()
    entry = auth.get(RM_USER) if isinstance(auth.get(RM_USER), dict) else {}
    entry['rm_pw_set_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    auth[RM_USER] = entry
    _save_auth(auth)


def _rm_pw_expired() -> bool:
    """RM 비밀번호가 만료 주기를 넘겼는지. 설정 시각이 없으면 만료 아님(최초 로그인 시 기록)."""
    ts = _rm_pw_set_at()
    if not ts:
        return False
    try:
        set_at = datetime.strptime(ts, '%Y-%m-%d %H:%M:%S')
    except (ValueError, TypeError):
        return False
    return (datetime.now() - set_at).days >= RM_PW_MAX_AGE_DAYS


def _rm_days_left():
    """RM 비밀번호 만료까지 남은 일수(연금컨설팅팀 화면 표시용). RM 미설정 시 None.
    시계 미시작(최초 로그인 전)이면 전체 기간으로 표시. 음수면 이미 만료."""
    if not RM_USER:
        return None
    ts = _rm_pw_set_at()
    if not ts:
        return RM_PW_MAX_AGE_DAYS
    try:
        set_at = datetime.strptime(ts, '%Y-%m-%d %H:%M:%S')
    except (ValueError, TypeError):
        return RM_PW_MAX_AGE_DAYS
    return RM_PW_MAX_AGE_DAYS - (datetime.now() - set_at).days


@app.before_request
def _require_login():
    if request.endpoint in _PUBLIC_ENDPOINTS:
        return
    if not session.get('logged_in'):
        # API(fetch) 요청은 401 JSON, 일반 페이지는 로그인 화면으로 유도
        if request.path.startswith('/api/'):
            return jsonify({'success': False, 'message': '로그인이 필요합니다',
                            'login_required': True}), 401
        return redirect(url_for('login', next=request.path))
    # 로그인됨 — RM(조회 전용)은 관리 엔드포인트 차단(서버측 강제)
    if session.get('role') == 'rm' and request.endpoint in _ADMIN_ONLY_ENDPOINTS:
        if request.path.startswith('/api/'):
            return jsonify({'success': False, 'message': '조회 전용 계정은 이 기능을 사용할 수 없습니다.'}), 403
        return redirect(url_for('index'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = ''
    if request.method == 'POST':
        u = request.form.get('username', '').strip()
        p = request.form.get('password', '')
        role = _role_for(u)
        if role and verify_password(u, p):
            if role == 'rm' and _rm_pw_expired():
                # 2주 경과 → RM 로그인 차단(팝업 안내), 연금컨설팅팀 재설정 필요
                return render_template('login.html', error='', rm_expired=True)
            if role == 'rm' and not _rm_pw_set_at():
                _mark_rm_pw_set_at()   # 초기 비밀번호 최초 로그인 → 만료 시계 시작
            session['logged_in'] = True
            session['user'] = u
            session['role'] = role     # consulting=전체관리, rm=조회·다운로드만
            # 조회 방식: simple=간편조회(모바일 v2 신규), web=웹버전조회(기존 v1). 기본 web.
            session['view'] = 'simple' if request.form.get('view') == 'simple' else 'web'
            nxt = request.args.get('next') or '/'
            if not nxt.startswith('/'):   # 오픈 리다이렉트 방지
                nxt = '/'
            return redirect(nxt)
        error = '아이디 또는 비밀번호가 올바르지 않습니다.'
    if session.get('logged_in'):
        return redirect(url_for('index'))
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/api/change_password', methods=['POST'])
def api_change_password():
    """비밀번호 변경. 해시만 저장하며 평문은 기록하지 않는다.

    규칙:
    - 연금컨설팅팀(consulting): 로그인 화면/로그인 상태에서 본인 비밀번호 변경(현재 비밀번호 확인).
      또한 로그인 상태에서 target=RM 지정 시 RM 비밀번호를 재설정(연금컨설팅팀 본인 비밀번호로 확인).
    - RM: 자기 비밀번호도 변경할 수 없다(연금컨설팅팀만 가능).
    """
    d = request.get_json(silent=True) or {}
    cur = d.get('current') or ''
    new = d.get('new') or ''
    confirm = d.get('confirm') or ''

    logged_in = bool(session.get('logged_in'))
    acting_user = session.get('user') if logged_in else (d.get('username') or '').strip()
    acting_role = session.get('role') if logged_in else _role_for(acting_user)
    # 변경 대상: target 지정 시 그 계정, 없으면 본인
    target = (d.get('target') or '').strip() or acting_user
    if not _role_for(acting_user) or not _role_for(target):
        return jsonify({'success': False, 'message': '아이디 또는 현재 비밀번호가 올바르지 않습니다.'}), 400
    target_role = _role_for(target)
    is_self = (target == acting_user)

    if target_role == 'rm':
        # RM 비밀번호는 '로그인한 연금컨설팅팀'만 변경 가능(연금컨설팅팀 본인 비밀번호로 확인)
        if not (logged_in and acting_role == 'consulting'):
            return jsonify({'success': False, 'message': 'RM 비밀번호는 연금컨설팅팀 계정으로 로그인 후에만 변경할 수 있습니다.'}), 403
        if not verify_password(acting_user, cur):
            return jsonify({'success': False, 'message': '연금컨설팅팀 비밀번호가 올바르지 않습니다.'}), 400
    else:
        # 연금컨설팅팀 대상: 본인만, 본인 현재 비밀번호 확인
        if not is_self:
            return jsonify({'success': False, 'message': '본인 비밀번호만 변경할 수 있습니다.'}), 403
        if logged_in and acting_role == 'rm':
            return jsonify({'success': False, 'message': 'RM 비밀번호는 연금컨설팅팀만 변경할 수 있습니다.'}), 403
        if not verify_password(target, cur):
            return jsonify({'success': False, 'message': '현재 비밀번호가 올바르지 않습니다.'}), 400

    if len(new) < MIN_PASSWORD_LEN:
        return jsonify({'success': False,
                        'message': f'새 비밀번호는 {MIN_PASSWORD_LEN}자 이상이어야 합니다.'}), 400
    if new != confirm:
        return jsonify({'success': False, 'message': '새 비밀번호가 서로 일치하지 않습니다.'}), 400
    if is_self and new == cur:
        return jsonify({'success': False, 'message': '현재 비밀번호와 다른 비밀번호를 입력하세요.'}), 400

    auth = _load_auth()
    # 구형(top-level password_hash)이 있으면 APP_USER 것으로 이관 후 신형 per-user로 저장
    if 'password_hash' in auth:
        migrated = {'password_hash': auth.pop('password_hash')}
        if 'updated' in auth:
            migrated['updated'] = auth.pop('updated')
        auth[APP_USER] = migrated
    entry = auth.get(target) if isinstance(auth.get(target), dict) else {}
    entry['password_hash'] = generate_password_hash(new)
    entry['updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if target_role == 'rm':
        entry['rm_pw_set_at'] = entry['updated']   # RM 만료 시계(2주) 리셋
    auth[target] = entry
    _save_auth(auth)
    logger.info('비밀번호 변경 완료 (대상=%s, 변경자=%s)', target, acting_user)
    return jsonify({'success': True, 'message': '비밀번호가 변경되었습니다.'})

# ── 백그라운드 조회 상태 ───────────────────────────────────────────────
_refresh_lock = threading.Lock()
_refresh_state: dict = {
    'running': False, 'total': 0, 'completed': 0,
    'results': {}, 'summary': '', 'updated_at': '', 'error': '',
    'started_at': 0.0,       # time.time() — 멈춘 조회 판단용
}
# 조회 스레드가 멈추면(크롬 응답없음 등) running 플래그가 영원히 True로 남아
# 이후 모든 조회가 '이미 조회 중'으로 막힌다. 이 시간이 지나면 죽은 것으로 보고 새 조회를 허용한다.
STALE_REFRESH_SEC = 30 * 60

RATINGS_FILE = os.path.join(DATA_DIR, 'ratings.json')
INSTITUTIONS_FILE = os.path.join(DATA_DIR, 'institutions.json')
EXPORT_TEMPLATE = os.path.join(BASE_DIR, 'export_template.xlsx')
OVERRIDES_FILE = os.path.join(DATA_DIR, 'overrides.json')
HISTORY_FILE = os.path.join(DATA_DIR, 'rating_history.json')

# 서버 포트: 클라우드(Cloud Run 등)는 환경변수 PORT를 지정함. 없으면 로컬 기본 5000.
PORT = int(os.environ.get('PORT', 5000))

RATING_SCALE = [
    'AAA', 'AA+', 'AA', 'AA-',
    'A+', 'A', 'A-',
    'BBB+', 'BBB', 'BBB-',
    'BB+', 'BB', 'BB-',
    'B+', 'B', 'B-',
    'CCC+', 'CCC', 'CCC-',
    'CC', 'C', 'D',
]

INSURANCE_CATEGORIES = {'손해보험', '생명보험'}
AGENCIES = ['nice', 'kr', 'kis']
AGENCY_LABELS = {'nice': '나이스신용평가', 'kr': '한국기업평가', 'kis': '한국신용평가'}
# 기존 등급을 지우기 전에 요구하는 '연속 빈값' 횟수. 1회 빈값은 일시적 스크래핑 실패로 보고
# 값을 유지한다(카운터는 ratings.json의 '{평가사}_miss'에 저장).
MISS_LIMIT = 2


# ── Rating helpers ────────────────────────────────────────────────────

def get_lowest_rating(ratings: list) -> str:
    valid = [r for r in ratings if r and r.strip() in RATING_SCALE]
    if not valid:
        return ''
    return max(valid, key=lambda r: RATING_SCALE.index(r))


def compare_ratings(old: str, new: str) -> str:
    """등급 변화 방향: 'up' | 'down' | 'same' | ''"""
    if not old or not new:
        return ''
    if old not in RATING_SCALE or new not in RATING_SCALE:
        return ''
    diff = RATING_SCALE.index(new) - RATING_SCALE.index(old)
    if diff < 0:
        return 'up'
    if diff > 0:
        return 'down'
    return 'same'


def get_rating_type_label(agency_type: str) -> str:
    labels = {
        'ICR': '기업신용등급',
        '회사채선순위': '회사채 선순위',
        'IFS': '보험지급능력',
        '': '-',
    }
    return labels.get(agency_type, agency_type)


# ── Data I/O ──────────────────────────────────────────────────────────

def load_institutions() -> dict:
    with open(INSTITUTIONS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_ratings() -> dict:
    if not os.path.exists(RATINGS_FILE):
        return {}
    with open(RATINGS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_ratings(data: dict):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(RATINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── 신용등급 변경이력 ──────────────────────────────────────────────────
# 조회 중 등급 변경이 감지되면 이력을 누적한다.
# 레코드: {name, agency, agency_label, prev, current, direction,
#          type_label, eval_date(변경일=평가일), detected_at(감지일시)}

def load_history() -> list:
    if not os.path.exists(HISTORY_FILE):
        return []
    try:
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data if isinstance(data, list) else []
    except Exception:
        return []


def append_history(records: list):
    """새 변경 레코드들을 이력 파일에 누적(append). 비어 있으면 아무것도 안 함."""
    if not records:
        return
    os.makedirs(DATA_DIR, exist_ok=True)
    hist = load_history()
    hist.extend(records)
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(hist, f, ensure_ascii=False, indent=2)


# ── 사용자 강제 정정(override) ─────────────────────────────────────────
# 재조회(스크래핑)가 사용자가 정정한 값을 다시 덮어쓰지 못하도록, 표시·저장 시점에 강제.
# 형식: { "기관명": { "kr": null } }  → null이면 미공시(빈값) 강제, 문자열이면 그 등급 강제.

def load_overrides() -> dict:
    if not os.path.exists(OVERRIDES_FILE):
        return {}
    try:
        with open(OVERRIDES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _apply_overrides(name: str, d: dict, overrides: dict | None = None) -> dict:
    """기관 데이터 dict(d)에 override를 in-place 적용. 정정값이 재조회에도 유지되게 함."""
    ov = (overrides if overrides is not None else load_overrides()).get(name)
    if not ov:
        return d
    for ag, val in ov.items():
        if ag not in AGENCIES:
            continue
        d[ag] = val or ''
        d[f'{ag}_eval_date'] = ''
        d[f'{ag}_type'] = ''
        d[f'{ag}_prev'] = ''
        d[f'{ag}_changed'] = False
    # override가 적용되면 최종등급 변경표시는 무의미 → 초기화
    d['final_prev'] = ''
    d['final_changed'] = False
    return d


# ── Build view data ───────────────────────────────────────────────────

def build_row(name: str, category: str, inst_data: dict, overrides: dict | None = None) -> dict:
    is_insurance = category in INSURANCE_CATEGORIES
    inst_data = _apply_overrides(name, dict(inst_data), overrides)

    agencies_out = {}
    any_changed = False

    for ag in AGENCIES:
        current = inst_data.get(ag, '')
        prev    = inst_data.get(f'{ag}_prev', '')
        changed = inst_data.get(f'{ag}_changed', False)
        direction = compare_ratings(prev, current) if changed and prev else ''
        raw_type = inst_data.get(f'{ag}_type', '')
        if is_insurance:
            raw_type = 'IFS'

        agencies_out[ag] = {
            'rating':    current,
            'eval_date': inst_data.get(f'{ag}_eval_date', ''),
            'prev':      prev,
            'changed':   changed,
            'direction': direction,
            'type':      raw_type,
            'type_label': get_rating_type_label(raw_type),
        }
        if changed and prev and current != prev:
            any_changed = True

    ratings_list = [agencies_out[ag]['rating'] for ag in AGENCIES]
    final = get_lowest_rating(ratings_list)

    final_prev = inst_data.get('final_prev', '')
    final_changed = inst_data.get('final_changed', False)
    final_direction = compare_ratings(final_prev, final) if final_changed and final_prev else ''

    # 최종등급 근거: 어느 평가사의 어떤 유형 등급인지
    if final:
        contrib_ags = [ag for ag in AGENCIES if agencies_out[ag]['rating'] == final]
        contrib_labels = [AGENCY_LABELS[ag] for ag in contrib_ags]
        contrib_types = list(dict.fromkeys(
            agencies_out[ag]['type_label'] for ag in contrib_ags if agencies_out[ag]['type']
        ))
        if len(contrib_ags) == 3:
            basis_agency = '3사 동일'
        elif len(contrib_ags) == 2:
            basis_agency = ' · '.join(contrib_labels)
        else:
            basis_agency = contrib_labels[0]
        basis_type = ' · '.join(contrib_types) if contrib_types else ''
    else:
        basis_agency = ''
        basis_type = ''

    # 등급 상이 여부: 등급이 있는 평가사들 사이에 서로 다른 등급이 존재하는지
    rated = [(ag, agencies_out[ag]['rating']) for ag in AGENCIES if agencies_out[ag]['rating']]
    rating_counts = Counter(r for _, r in rated)
    rating_mismatch = len(rating_counts) > 1
    mismatch_agencies = []
    if rating_mismatch:
        if all(c == 1 for c in rating_counts.values()):
            # 모든 평가사 등급이 제각각 → 전부 상이 처리
            mismatch_agencies = [AGENCY_LABELS[ag] for ag, _ in rated]
        else:
            top_count = max(rating_counts.values())
            majority = {r for r, c in rating_counts.items() if c == top_count}
            mismatch_agencies = [AGENCY_LABELS[ag] for ag, r in rated if r not in majority]

    # 변경사항 목록
    changes = []
    for ag in AGENCIES:
        ag_d = agencies_out[ag]
        if ag_d['changed'] and ag_d['prev'] and ag_d['rating'] and ag_d['rating'] != ag_d['prev']:
            changes.append({
                'agency_label': AGENCY_LABELS[ag],
                'prev': ag_d['prev'],
                'current': ag_d['rating'],
                'direction': ag_d['direction'],
                'type_label': ag_d['type_label'],
                'eval_date': ag_d['eval_date'],
            })

    # 등급구분 대표값
    types_used = [agencies_out[ag]['type'] for ag in AGENCIES if agencies_out[ag]['rating']]
    if is_insurance:
        rep_type = 'IFS'
    elif types_used and all(t == '회사채선순위' for t in types_used):
        rep_type = '회사채선순위'
    elif types_used and all(t == 'ICR' for t in types_used):
        rep_type = 'ICR'
    elif '회사채선순위' in types_used:
        rep_type = 'ICR/회사채혼용'
    else:
        rep_type = ''

    return {
        'name': name,
        'category': category,
        'is_insurance': is_insurance,
        'agencies': agencies_out,
        'final': final,
        'final_prev': final_prev,
        'final_changed': final_changed,
        'final_direction': final_direction,
        'basis_agency': basis_agency,
        'basis_type': basis_type,
        'rating_mismatch': rating_mismatch,
        'mismatch_agencies': mismatch_agencies,
        'changes': changes,
        'any_changed': any_changed,
        'scrape_status': inst_data.get('scrape_status', ''),
        'updated': inst_data.get('updated', ''),
        'rep_type': rep_type,
        'rep_type_label': get_rating_type_label(rep_type),
    }


def build_response_data(institutions: dict, ratings: dict) -> dict:
    category_order = ['증권', '시중은행', '지방은행', '저축은행', '손해보험', '생명보험', '기타']
    overrides = load_overrides()
    result = {}
    for category in category_order:
        items = institutions.get(category, [])
        rows = []
        for inst in items:
            name = inst['name']
            inst_data = ratings.get(name, {})
            rows.append(build_row(name, category, inst_data, overrides))
        result[category] = rows
    return result


# ── Template globals ──────────────────────────────────────────────────

def _rating_css(r: str) -> str:
    if r == 'AAA':        return 'rating-AAA'
    if r.startswith('AA'): return 'rating-AA'
    if r.startswith('A'):  return 'rating-A'
    if r.startswith('BBB'): return 'rating-BBB'
    if r.startswith('BB'): return 'rating-BB'
    if r.startswith('B'):  return 'rating-B'
    if r.startswith('CCC'): return 'rating-CCC'
    if r in ('CC', 'C', 'D'): return 'rating-low'
    return ''


@app.template_global()
def rating_badge(r: str, extra_cls: str = '') -> Markup:
    if not r:
        return Markup('<span class="no-rating">—</span>')
    css = _rating_css(r)
    return Markup(f'<span class="rbadge {css} {extra_cls}">{r}</span>')


@app.template_global()
def change_html(prev: str, current: str, direction: str) -> Markup:
    if not prev or not current or prev == current:
        return Markup('')
    arrow = {'up': '▲', 'down': '▼', 'same': '→'}.get(direction, '→')
    cls   = {'up': 'chg-up', 'down': 'chg-down', 'same': 'chg-same'}.get(direction, '')
    return Markup(
        f'<span class="chg-tag {cls}">'
        f'{prev}&nbsp;{arrow}&nbsp;{current}'
        f'</span>'
    )


@app.template_global()
def rating_select_options() -> Markup:
    opts = '<option value="">— 미공시 —</option>'
    for r in RATING_SCALE:
        opts += f'<option value="{r}">{r}</option>'
    return Markup(opts)


# ── Routes ────────────────────────────────────────────────────────────

@app.route('/pension')
def pension():
    """원리금보장상품 금리관리 화면(통합 탭). 순수 HTML을 그대로 서빙(Jinja 미처리)."""
    return send_file(os.path.join(BASE_DIR, 'pension.html'))


# ── 퇴직연금 금리 데이터 서버 저장(모든 기기/브라우저 공유) ──
_PENSION_STORE = os.path.join(BASE_DIR, 'data', 'pension_store.json')


@app.route('/api/pension_store', methods=['GET'])
def pension_store_get():
    """저장된 월별 금리 데이터(DB) 반환."""
    if os.path.exists(_PENSION_STORE):
        with open(_PENSION_STORE, encoding='utf-8') as f:
            return app.response_class(f.read(), mimetype='application/json')
    return jsonify({})


@app.route('/api/pension_store', methods=['POST'])
def pension_store_post():
    """월별 금리 데이터(DB)를 서버에 저장 → 다른 기기에서도 조회 가능."""
    data = request.get_json(force=True, silent=True)
    if not isinstance(data, dict):
        return jsonify({'error': 'invalid'}), 400
    os.makedirs(os.path.dirname(_PENSION_STORE), exist_ok=True)
    with open(_PENSION_STORE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)
    return jsonify({'ok': True, 'months': sorted(data.keys())})


# ── 상품제안관리 ───────────────────────────────────────────────────────
#   proposals.json      : 기준월 -> '이달의 제안상품' 목록(등록된 것)
#   proposal_meta.json  : 상품키 -> {universe:'Y'|'N', sellable:bool}
#     (유니버스·판매가능여부는 원리금 데이터에 없어 이 화면에서 관리한다)
#   상품키 = "sector|org|fam|months" (proposal.html의 rowKey()와 동일 규칙)
_PROPOSALS_FILE = os.path.join(BASE_DIR, 'data', 'proposals.json')
_PROPOSAL_META_FILE = os.path.join(BASE_DIR, 'data', 'proposal_meta.json')


def _load_json_file(path: str) -> dict:
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        logger.exception('%s 읽기 실패', os.path.basename(path))
        return {}


def _save_json_file(path: str, data: dict):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


@app.route('/proposal')
def proposal():
    """상품제안관리 화면(통합 탭). 순수 HTML 서빙(Jinja 미처리)."""
    return send_file(os.path.join(BASE_DIR, 'proposal.html'))


@app.route('/api/proposals', methods=['GET'])
def proposals_get():
    """기준월별 '이달의 제안상품' 목록 전체 반환."""
    return jsonify(_load_json_file(_PROPOSALS_FILE))


@app.route('/api/proposals', methods=['POST'])
def proposals_post():
    """특정 기준월의 제안상품 목록 저장(등록/수정). payload: {month, items:[...]}"""
    d = request.get_json(force=True, silent=True) or {}
    month = (d.get('month') or '').strip()
    items = d.get('items')
    if not month or not isinstance(items, list):
        return jsonify({'success': False, 'message': 'month/items 형식 오류'}), 400
    store = _load_json_file(_PROPOSALS_FILE)
    store[month] = items
    _save_json_file(_PROPOSALS_FILE, store)
    return jsonify({'success': True, 'month': month, 'count': len(items)})


@app.route('/api/proposal_meta', methods=['GET'])
def proposal_meta_get():
    """상품키별 유니버스/판매가능 메타 반환."""
    return jsonify(_load_json_file(_PROPOSAL_META_FILE))


@app.route('/api/proposal_meta', methods=['POST'])
def proposal_meta_post():
    """상품키별 메타 저장. payload: {key, universe:'Y'|'N', precheck:'Y'|'N', sellable:bool}"""
    d = request.get_json(force=True, silent=True) or {}
    key = (d.get('key') or '').strip()
    if not key:
        return jsonify({'success': False, 'message': 'key 없음'}), 400
    meta = _load_json_file(_PROPOSAL_META_FILE)
    entry = meta.get(key, {})
    if 'universe' in d:
        entry['universe'] = 'N' if d['universe'] == 'N' else 'Y'
    if 'precheck' in d:                       # 연컨사전확인 Y/N
        entry['precheck'] = 'Y' if d['precheck'] == 'Y' else 'N'
    if 'sellable' in d:
        entry['sellable'] = bool(d['sellable'])
    meta[key] = entry
    _save_json_file(_PROPOSAL_META_FILE, meta)
    return jsonify({'success': True})


# ── 공지사항 ───────────────────────────────────────────────────────────
#   연금컨설팅팀이 상품제안관리에서 입력 → 모바일 제안카드 「공지」 버튼 팝업에 표시.
#   data/notice.json = {"text": "...", "updated_at": "YYYY-MM-DD HH:MM:SS"}. 런타임 데이터 → 배포 제외.
_NOTICE_FILE = os.path.join(BASE_DIR, 'data', 'notice.json')


@app.route('/api/notice', methods=['GET'])
def api_notice_get():
    """공지 내용 반환(로그인한 모든 사용자)."""
    d = _load_json_file(_NOTICE_FILE)
    return jsonify({'text': d.get('text', ''), 'updated_at': d.get('updated_at', '')})


@app.route('/api/notice', methods=['POST'])
def api_notice_post():
    """공지 저장(연금컨설팅팀 전용 — _ADMIN_ONLY_ENDPOINTS)."""
    d = request.get_json(force=True, silent=True) or {}
    text = d.get('text', '')
    if not isinstance(text, str):
        return jsonify({'success': False, 'message': 'text 형식 오류'}), 400
    text = text[:2000]   # 과도한 길이 방지
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    _save_json_file(_NOTICE_FILE, {'text': text, 'updated_at': now})
    return jsonify({'success': True, 'updated_at': now})


# ── 증권사 ELB/DLB 발행정지기간 ───────────────────────────────────────
#   연금컨설팅팀이 상품제안관리에서 기관별 입력 → 모바일 상세 팝업에 표시.
#   data/issue_suspension.json = {"삼성증권㈜": {"start": "2026-08-20", "end": "2026-08-26"}, ...}
#   키 = 기관명(org). 그 증권사의 모든 ELB/DLB에 적용. 런타임 데이터 → 배포 제외.
_ISSUE_SUSPENSION_FILE = os.path.join(BASE_DIR, 'data', 'issue_suspension.json')


@app.route('/api/issue_suspension', methods=['GET'])
def api_issue_suspension_get():
    """기관별 발행정지기간 반환(로그인한 모든 사용자)."""
    return jsonify(_load_json_file(_ISSUE_SUSPENSION_FILE))


@app.route('/api/issue_suspension', methods=['POST'])
def api_issue_suspension_post():
    """발행정지기간 전체 맵 저장(연금컨설팅팀 전용). payload: {map: {org:{start,end}}}"""
    d = request.get_json(force=True, silent=True) or {}
    m = d.get('map')
    if not isinstance(m, dict):
        return jsonify({'success': False, 'message': 'map 형식 오류'}), 400
    clean = {}
    for org, v in m.items():
        if isinstance(v, dict):
            s = str(v.get('start', ''))[:10]
            e = str(v.get('end', ''))[:10]
            if s or e:                       # 둘 다 비면 미저장(=해제)
                clean[str(org)[:100]] = {'start': s, 'end': e}
    _save_json_file(_ISSUE_SUSPENSION_FILE, clean)
    return jsonify({'success': True, 'count': len(clean)})


# ── 과거 금리 추이(내부 보관 데이터) ──
#   data/rate_history.xlsx : 원본 엑셀 바이트 그대로 보관(다운로드용)
#   data/rate_history.json : 구조화 테이블(향후 개발에서 재사용)
_RATE_HISTORY_XLSX = os.path.join(BASE_DIR, 'data', 'rate_history.xlsx')
_RATE_HISTORY_JSON = os.path.join(BASE_DIR, 'data', 'rate_history.json')


@app.route('/download/rate_history')
def download_rate_history():
    """'과거 금리 추이' 버튼 → 내부 보관된 원본 엑셀을 그대로 다운로드."""
    if not os.path.exists(_RATE_HISTORY_XLSX):
        return jsonify({'error': 'rate_history.xlsx not found'}), 404
    return send_file(_RATE_HISTORY_XLSX, as_attachment=True,
                     download_name='과거 금리 추이.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/rate_history')
def api_rate_history():
    """과거 금리 추이 구조화 테이블(JSON). 향후 개발에서 재사용."""
    if not os.path.exists(_RATE_HISTORY_JSON):
        return jsonify({'error': 'rate_history.json not found'}), 404
    with open(_RATE_HISTORY_JSON, encoding='utf-8') as f:
        return app.response_class(f.read(), mimetype='application/json')


@app.route('/download/rate_compare')
def download_rate_compare():
    """'금리비교하기' 버튼 → 과거 금리 추이 데이터로 5개 꺾은선 그래프가 든 엑셀 생성.
       완성본 Sheet2 사용(Sheet1은 #DIV/0! 다수로 값 누락). 컬럼: 1 DATE,
       2 증권평균, 3 은행평균, 4 생보평균, 5 손보평균, 6 저축평균, 7 원리금보장DB, 8 기준금리."""
    import openpyxl
    from collections import defaultdict
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.axis import ChartLines
    from openpyxl.chart.text import RichText
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    from openpyxl.drawing.text import (Paragraph, ParagraphProperties,
                                       CharacterProperties, RichTextProperties)
    if not os.path.exists(_RATE_HISTORY_JSON):
        return jsonify({'error': 'rate_history.json not found'}), 404
    with open(_RATE_HISTORY_JSON, encoding='utf-8') as f:
        hist = json.load(f)
    sh = hist['sheets']['Sheet2']
    headers = sh['headers']
    rows = sh['rows']

    NCOL = 8  # DATE ~ 기준금리 (Sheet2 인덱스 0~7)

    def num(v):
        if isinstance(v, bool):
            return None
        if isinstance(v, (int, float)):
            return round(v, 2)  # 금리 소수점 둘째자리
        try:
            return round(float(v), 2)
        except (TypeError, ValueError):
            return None  # '#DIV/0!' · 빈값 등은 공백 처리

    wb = openpyxl.Workbook()
    wsd = wb.active
    wsd.title = '데이터'
    wsd.append(headers[:NCOL])
    for r in rows:
        wsd.append([r[0]] + [num(r[i]) for i in range(1, NCOL)])
    last = wsd.max_row  # 헤더 포함 마지막 행
    cats = Reference(wsd, min_col=1, min_row=2, max_row=last)

    def x_label_style():
        """X축 날짜 라벨: 세로(-90°) 회전 + 작은 글꼴 → 겹침 없이 식별 가능."""
        return RichText(
            bodyPr=RichTextProperties(rot=-5400000, vert='horz'),
            p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=700)))])

    # 다중 라인 색상 순서: 주황 → 네이비 → 하늘색 → 짙은 검정 (그 뒤 회색·진주황)
    LINE_COLORS = ['F68121', '123E7C', '5AB0E0', '1A1A1A', '9AA3AE', 'C77B2E']

    def style_series(ch, marker_size):
        """색상순서 라인. marker_size>0이면 원형 마커(흰 채움·색 테두리), 0이면 마커 없음(깔끔한 꺾은선)."""
        for i, s in enumerate(ch.series):
            col = LINE_COLORS[i % len(LINE_COLORS)]
            gp = GraphicalProperties()
            gp.line = LineProperties(solidFill=col, w=25400)   # 2pt
            s.graphicalProperties = gp
            if marker_size and marker_size > 0:
                mk = Marker(symbol='circle', size=marker_size)
                mgp = GraphicalProperties(solidFill='FFFFFF')
                mgp.line = LineProperties(solidFill=col)
                mk.graphicalProperties = mgp
                s.marker = mk
            else:
                s.marker = Marker(symbol='none')   # ①~⑤: 동그라미 제거
            s.smooth = False

    def make_chart(ws, cats_ref, last_row, title, cols, legend, marker_size, xrot):
        ch = LineChart()
        if title:
            ch.title = title
        ch.type = 'line'
        ch.style = 2
        ch.height = 13
        ch.width = 30
        ch.y_axis.delete = False
        ch.x_axis.delete = False
        ch.y_axis.numFmt = '0.00'
        ch.y_axis.majorGridlines = ChartLines()
        if xrot:
            ch.x_axis.txPr = x_label_style()    # 반월 시계열: 날짜 라벨 세로 회전
        ch.x_axis.tickLblPos = 'low'
        for c in cols:
            ref = Reference(ws, min_col=c, min_row=1, max_row=last_row)
            ch.add_data(ref, titles_from_data=True)
        ch.set_categories(cats_ref)
        style_series(ch, marker_size)
        if legend:
            ch.legend.position = legend
        else:
            ch.legend = None
        return ch

    # 반월 시계열 5종 (마커 작게, 상단범례/단일=제목)
    specs = [
        (None, '①업권평균비교', [2, 3, 4, 5, 6, 7], 't'),
        ('증권사 ELB', '②증권ELB', [2], None),
        ('원리금보장상품 금리 평균 (1년, DB)', '③원리금보장평균', [7], None),
        (None, '④증권vs기준금리', [2, 8], 't'),
        (None, '⑤업권별사업자금리', [2, 3, 4], 't'),
    ]
    for title, sheet_name, cols, legend in specs:
        wsc = wb.create_sheet(title=sheet_name)
        wsc.add_chart(make_chart(wsd, cats, last, title, cols, legend, 0, True), 'B2')

    # ⑥ 원리금보장 평균금리 vs 소비자물가상승률 (연도별 2015~, 첨부 그래프 형식)
    ann = defaultdict(list)
    for r in rows:
        y = str(r[0])[:4]
        v = num(r[6])   # 원리금보장상품금리 평균(1년, DB)
        if v is not None and y.isdigit():
            ann[y].append(v)
    rate_annual = {y: round(sum(vs) / len(vs), 2) for y, vs in ann.items()}
    cpi_annual = {}
    cd = _cpi_history_load()
    if cd:
        cpi_annual.update(cd.get('annual', {}))
        monthly = cd.get('monthly', {})
        if monthly:
            cy = sorted(set(k[:4] for k in monthly))[-1]
            vals = [v for k, v in monthly.items() if k[:4] == cy]
            cpi_annual[cy] = round(sum(vals) / len(vals), 2)
    wsd2 = wb.create_sheet(title='데이터_연도별')
    wsd2.append(['연도', '원리금보장상품 평균금리', '소비자물가상승률'])
    for y in range(2015, 2027):
        ys = str(y)
        wsd2.append([ys, rate_annual.get(ys), cpi_annual.get(ys)])
    last2 = wsd2.max_row
    cats2 = Reference(wsd2, min_col=1, min_row=2, max_row=last2)
    wsc6 = wb.create_sheet(title='⑥원리금보장vs물가상승률')
    #  범례는 상단('t') — 하단이면 연도 라벨과 겹침(참조 이미지=상단)
    wsc6.add_chart(make_chart(wsd2, cats2, last2, None, [2, 3], 't', 7, False), 'B2')

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name='금리 비교.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 소비자물가지수(e-나라지표) 최신월 상승률 ──
_CPI_CACHE = {'ts': 0, 'data': None}
_CPI_URL = ('https://www.index.go.kr/unity/potal/eNara/sub/showStblGams3.do'
            '?stts_cd=106001&idx_cd=1060&freq=M&period=N')

# ── 과거 물가상승률(내부 보관: 연도별 + 올해 월별 누적) ──
_CPI_HISTORY_JSON = os.path.join(BASE_DIR, 'data', 'cpi_history.json')


def _cpi_history_load():
    if not os.path.exists(_CPI_HISTORY_JSON):
        return None
    with open(_CPI_HISTORY_JSON, encoding='utf-8') as f:
        return json.load(f)


def _cpi_history_save(d):
    with open(_CPI_HISTORY_JSON, 'w', encoding='utf-8') as f:
        json.dump(d, f, ensure_ascii=False, indent=1)


def _cpi_upsert_month(year, month, rate):
    """최신월 물가상승률을 과거 물가상승률 테이블(월별)에 누적 저장(매월 자동)."""
    try:
        d = _cpi_history_load()
        if not d:
            return
        key = '%s-%02d' % (str(year), int(month))
        d.setdefault('monthly', {})
        if d['monthly'].get(key) != rate:
            d['monthly'][key] = rate
            _cpi_history_save(d)
    except Exception:  # noqa
        logger.exception('물가상승률 월별 누적 저장 실패')


def _cpi_graph_points(d):
    """연도별(과거) + 올해(1월~최신월 평균) 그래프 포인트."""
    annual = d.get('annual', {})
    monthly = d.get('monthly', {})
    pts = [{'x': y, 'y': annual[y]} for y in sorted(annual)]
    if monthly:
        cy = sorted(set(k[:4] for k in monthly))[-1]     # 올해(월별 최신 연도)
        vals = [v for k, v in monthly.items() if k[:4] == cy]
        pts.append({'x': cy, 'y': round(sum(vals) / len(vals), 2),
                    'avg': True, 'months': len(vals)})
    return pts


# ── 시장금리 4종(국고채1·3년·회사채AA-·CD91일) — 주요지표 순환 카드용 ──
#   data/bond_rates.json 을 매일 갱신(스케줄러/스크래퍼). 파일 없으면 기본값 반환.
BOND_RATES_FILE = os.path.join(DATA_DIR, 'bond_rates.json')
_BOND_DEFAULT = {'date': '2026-08-20',
                 'rates': {'ktb1': 3.404, 'ktb3': 3.811, 'corpAA': 4.504, 'cd91': 2.90},
                 'updated': ''}


@app.route('/api/bond_rates')
def api_bond_rates():
    try:
        with open(BOND_RATES_FILE, encoding='utf-8') as f:
            return jsonify(json.load(f))
    except Exception:
        return jsonify(_BOND_DEFAULT)


ECOS_KEY_FILE = os.path.join(BASE_DIR, 'ecos_key.txt')


def _ecos_key():
    k = (os.environ.get('ECOS_API_KEY') or '').strip()
    if k:
        return k
    try:
        with open(ECOS_KEY_FILE, encoding='utf-8') as f:
            return f.read().strip()
    except Exception:
        return ''


def fetch_bond_rates():
    """ECOS 817Y002(시장금리 일별)에서 국고채1·3년/회사채AA-/CD91일 최신값 → bond_rates.json 갱신.
    항목은 코드가 아니라 ITEM_NAME1(국고채+1년 등)으로 매칭해 코드 변경에 견고하게 처리."""
    import datetime as _dt
    import requests as _rq
    key = _ecos_key()
    if not key:
        logger.warning('ECOS_API_KEY 미설정 — 시장금리 갱신 건너뜀(ecos_key.txt 또는 환경변수)')
        return False
    end = _dt.date.today()
    start = end - _dt.timedelta(days=14)
    url = ('https://ecos.bok.or.kr/api/StatisticSearch/%s/json/kr/1/700/817Y002/D/%s/%s'
           % (key, start.strftime('%Y%m%d'), end.strftime('%Y%m%d')))
    try:
        j = _rq.get(url, timeout=20).json()
    except Exception:
        logger.exception('ECOS 요청 실패')
        return False
    rows = (j.get('StatisticSearch') or {}).get('row') or []
    if not rows:
        logger.warning('ECOS 응답에 데이터 없음(키/파라미터 확인): %s', str(j)[:200])
        return False

    def pick(pred):
        cand = []
        for x in rows:
            if pred(x.get('ITEM_NAME1', '')):
                try:
                    cand.append((x.get('TIME', ''), float(x['DATA_VALUE'])))
                except Exception:
                    continue
        if not cand:
            return None
        cand.sort()
        return round(cand[-1][1], 3), cand[-1][0]

    targets = {
        'ktb1':   lambda n: '국고채' in n and '1년' in n,
        'ktb3':   lambda n: '국고채' in n and '3년' in n,
        'corpAA': lambda n: '회사채' in n and 'AA-' in n and '민평' not in n,  # KOFIA 최종호가(민간평가 제외)
        'cd91':   lambda n: 'CD' in n and '91' in n,
    }
    rates, latest = {}, ''
    for k, pred in targets.items():
        got = pick(pred)
        if got:
            rates[k] = got[0]
            latest = max(latest, got[1])
    if not rates:
        logger.warning('ECOS 항목 매칭 실패(ITEM_NAME1 확인)')
        return False
    date_str = '%s-%s-%s' % (latest[:4], latest[4:6], latest[6:8]) if len(latest) == 8 else ''
    out = {'date': date_str, 'rates': rates,
           'updated': _dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    try:
        with open(BOND_RATES_FILE, 'w', encoding='utf-8') as f:
            json.dump(out, f, ensure_ascii=False)
        logger.info('시장금리 갱신 완료: %s', out)
        return True
    except Exception:
        logger.exception('bond_rates 저장 실패')
        return False


@app.route('/admin/refresh_bond_rates', methods=['POST'])
def admin_refresh_bond_rates():
    if fetch_bond_rates():
        try:
            with open(BOND_RATES_FILE, encoding='utf-8') as f:
                return jsonify({'success': True, 'data': json.load(f)})
        except Exception:
            return jsonify({'success': True})
    return jsonify({'success': False, 'message': 'ECOS 갱신 실패 — 서버 로그/키 확인'}), 500


# ── 약관·상품설명서 관리 (상품제공기관별 PDF · 종류별 다중 파일) ─────────
DOC_DIR = os.path.join(BASE_DIR, 'doc_files')
DOCS_META_FILE = os.path.join(DATA_DIR, 'docs_meta.json')
DOC_KINDS = ('약관', '상품설명서')
# ZIP/업로드 기관명 → 시스템 등록명 별칭
_DOC_ALIAS = {
    'SC은행': 'SC제일은행', 'DB저축은행': '디비저축은행', 'IBK저축은행': '아이비케이저축은행',
    'JT친애저축은행': '제이티친애저축은행', 'NH저축은행': '엔에이치저축은행',
    'IBK연금보험': '아이비케이연금보험', '우정사업본부': '우체국',
}
# 신용등급 목록엔 없지만 상품설명서용으로 추가하는 상품제공기관
_EXTRA_DOC_ORGS = [('증권', '우리투자증권'), ('기타', '우체국')]


def _load_docs_meta():
    try:
        with open(DOCS_META_FILE, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _save_docs_meta(m):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(DOCS_META_FILE, 'w', encoding='utf-8') as f:
        json.dump(m, f, ensure_ascii=False, indent=2)


def _org_key(name):
    return re.sub(r'[\\/:*?"<>|]+', '_', str(name)).strip() or '_'


def _safe_fname(fn):
    fn = re.sub(r'[\\/:*?"<>|]+', '_', os.path.basename(str(fn))).strip()
    return fn or 'file.pdf'


def _all_institutions():
    out = []
    for cat, items in load_institutions().items():
        for it in items:
            out.append((cat, it['name'] if isinstance(it, dict) else it))
    have = {n for _, n in out}
    for cat, nm in _EXTRA_DOC_ORGS:
        if nm not in have:
            out.append((cat, nm))
            have.add(nm)
    return out


def _doc_add(meta, org, kind, orig_filename, data_bytes):
    """org/kind 리스트에 파일 추가(동일 저장명은 교체). 실제 파일도 저장."""
    d = os.path.join(DOC_DIR, _org_key(org), kind)
    os.makedirs(d, exist_ok=True)
    stored = _safe_fname(orig_filename)
    with open(os.path.join(d, stored), 'wb') as out:
        out.write(data_bytes)
    lst = meta.setdefault(org, {}).setdefault(kind, [])
    lst[:] = [x for x in lst if x.get('stored') != stored]
    lst.append({'filename': os.path.basename(orig_filename), 'stored': stored,
                'uploaded': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})


@app.route('/docs')
def docs_page():
    return send_file(os.path.join(BASE_DIR, 'docs.html'))


@app.route('/docs_mobile')
def docs_mobile_page():
    return send_file(os.path.join(BASE_DIR, 'docs_mobile.html'))


@app.route('/api/docs')
def api_docs_list():
    meta = _load_docs_meta()
    items, reg = [], {'약관': 0, '상품설명서': 0}
    for cat, nm in _all_institutions():
        e = meta.get(nm, {})
        row = {'org': nm, 'cat': cat}
        for k in DOC_KINDS:
            files = e.get(k) or []
            row[k] = [{'filename': x.get('filename'), 'stored': x.get('stored')} for x in files]
            if files:
                reg[k] += 1
        items.append(row)
    return jsonify({'total': len(items), 'registered': reg, 'items': items})


@app.route('/api/docs/upload', methods=['POST'])
def api_docs_upload():
    org = (request.form.get('org') or '').strip()
    kind = (request.form.get('kind') or '').strip()
    f = request.files.get('file')
    if not org or kind not in DOC_KINDS or not f or not f.filename:
        return jsonify({'success': False, 'message': '잘못된 요청'}), 400
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({'success': False, 'message': 'PDF 파일만 업로드할 수 있습니다'}), 400
    meta = _load_docs_meta()
    _doc_add(meta, org, kind, f.filename, f.read())
    _save_docs_meta(meta)
    return jsonify({'success': True, 'filename': f.filename})


@app.route('/api/docs/download')
def api_docs_download():
    org = (request.args.get('org') or '').strip()
    kind = (request.args.get('kind') or '').strip()
    stored = _safe_fname(request.args.get('stored') or '')
    if kind not in DOC_KINDS:
        return jsonify({'success': False, 'message': '잘못된 요청'}), 400
    files = _load_docs_meta().get(org, {}).get(kind) or []
    info = next((x for x in files if x.get('stored') == stored), None)
    path = os.path.join(DOC_DIR, _org_key(org), kind, stored)
    if not info or not os.path.exists(path):
        return jsonify({'success': False, 'message': '파일이 없습니다'}), 404
    return send_file(path, as_attachment=True, download_name=info.get('filename') or stored)


@app.route('/api/docs/delete', methods=['POST'])
def api_docs_delete():
    d = request.get_json(silent=True) or {}
    org = (d.get('org') or '').strip()
    kind = (d.get('kind') or '').strip()
    stored = _safe_fname(d.get('stored') or '')
    if kind not in DOC_KINDS:
        return jsonify({'success': False}), 400
    try:
        p = os.path.join(DOC_DIR, _org_key(org), kind, stored)
        if os.path.exists(p):
            os.remove(p)
    except Exception:
        logger.exception('문서 삭제 실패')
    meta = _load_docs_meta()
    lst = (meta.get(org, {}) or {}).get(kind)
    if lst:
        meta[org][kind] = [x for x in lst if x.get('stored') != stored]
        if not meta[org][kind]:
            del meta[org][kind]
        if not meta[org]:
            del meta[org]
        _save_docs_meta(meta)
    return jsonify({'success': True})


def _zip_name(zi):
    if zi.flag_bits & 0x800:
        return zi.filename
    try:
        return zi.filename.encode('cp437').decode('cp949')
    except Exception:
        return zi.filename


def _resolve_org(name, orgs):
    # 1) 별칭(ZIP 짧은이름 → 시스템 정식명)
    for zipname, sysname in _DOC_ALIAS.items():
        if zipname in name:
            return sysname
    # 2) 시스템 정식명이 파일경로에 그대로 포함
    hit = next((o for o in orgs if o in name), None)
    if hit:
        return hit
    # 3) 대괄호/괄호 안 기관명 추출 → 공백 제거 후 양방향 부분매칭
    m = re.search(r'[\[\(]\s*([^\]\)]+?)\s*[\]\)]', name)
    if m:
        bn = m.group(1).replace(' ', '')
        for o in orgs:
            on = o.replace(' ', '')
            if bn and (bn in on or on in bn):
                return o
    return None


def classify_zip(data_bytes):
    """ZIP 바이트 → 기관/종류 자동 분류 후 저장. (matched, unmatched) 반환.
    HTTP 업로드(용량제한)와 무관하게 서버에서 직접 호출할 수도 있음(classify_zip.py)."""
    import zipfile
    import io as _io
    orgs = sorted([nm for _, nm in _all_institutions()], key=len, reverse=True)
    meta = _load_docs_meta()
    matched, unmatched = [], []
    z = zipfile.ZipFile(_io.BytesIO(data_bytes))
    for zi in z.infolist():
        try:
            if zi.is_dir():
                continue
            name = _zip_name(zi)
            base = os.path.basename(name)
            if not base.lower().endswith('.pdf'):
                continue
            if '설명서' in name or '셜명서' in name:      # 오타(셜명서) 보정
                kind = '상품설명서'
            elif ('약관' in name or '특약' in name):
                kind = '약관'
            else:
                unmatched.append(base)
                continue
            org = _resolve_org(name, orgs)
            if not org:
                unmatched.append(base)
                continue
            # SC제일은행: 약관에는 '특약'만 등록(기본약관·거치식약관 제외)
            if org == 'SC제일은행' and kind == '약관' and '특약' not in name:
                continue
            _doc_add(meta, org, kind, base, z.read(zi))
            matched.append(org + ' · ' + kind + ' · ' + base)
        except Exception:
            logger.exception('ZIP 파일 처리 실패: %s', getattr(zi, 'filename', ''))
            unmatched.append(getattr(zi, 'filename', '') or '?')
    _save_docs_meta(meta)
    return matched, unmatched


@app.route('/api/docs/bulk', methods=['POST'])
def api_docs_bulk():
    f = request.files.get('file')
    if not f or not (f.filename or '').lower().endswith('.zip'):
        return jsonify({'success': False, 'message': 'ZIP 파일을 올려주세요'}), 400
    try:
        matched, unmatched = classify_zip(f.read())
    except Exception as e:
        logger.exception('일괄 업로드 처리 실패')
        return jsonify({'success': False, 'message': '처리 실패: %s' % e}), 500
    return jsonify({'success': True, 'count': len(matched),
                    'matched': matched, 'unmatched': unmatched})


@app.route('/api/cpi')
def cpi_rate():
    """지표누리 e-나라지표 소비자물가지수에서 최신월 소비자물가 상승률(전년동월비)과 전월대비 변동을 반환."""
    import time
    now = time.time()
    if _CPI_CACHE['data'] and now - _CPI_CACHE['ts'] < 6 * 3600:
        return jsonify(_CPI_CACHE['data'])
    try:
        import requests
        from bs4 import BeautifulSoup
        html = requests.get(_CPI_URL, timeout=12, headers={'User-Agent': 'Mozilla/5.0'}).text
        soup = BeautifulSoup(html, 'lxml')
        target = next((t for t in soup.find_all('table') if '소비자물가' in t.get_text()), None)
        heads = [c.get_text(strip=True) for c in target.select('thead th, thead td')]
        row = None
        for tr in target.select('tbody tr'):
            cells = [c.get_text(strip=True) for c in tr.find_all(['th', 'td'])]
            if cells and cells[0].replace(' ', '') == '소비자물가':
                row = cells
                break
        months = [(i, h) for i, h in enumerate(heads) if re.match(r'^\d{6}월$', h)]
        li = months[-1][0]
        rate = float(row[li])
        pi = months[-2][0] if len(months) >= 2 else li - 1
        prev = float(row[pi])
        data = {'ok': True, 'year': months[-1][1][:4], 'month': months[-1][1][4:6],
                'rate': rate, 'prev': prev, 'diff': round(rate - prev, 2),
                'source': 'e-나라지표 소비자물가지수'}
        _cpi_upsert_month(data['year'], data['month'], data['rate'])   # 매월 자동 누적
        _CPI_CACHE['ts'] = now
        _CPI_CACHE['data'] = data
        return jsonify(data)
    except Exception as e:  # noqa
        logger.exception('소비자물가지수 조회 실패')
        if _CPI_CACHE['data']:
            return jsonify(_CPI_CACHE['data'])
        return jsonify({'ok': False, 'message': str(e)})


@app.route('/api/cpi_history')
def api_cpi_history():
    """과거 물가상승률 그래프 데이터(연도별 + 올해 월평균)."""
    d = _cpi_history_load()
    if not d:
        return jsonify({'error': 'cpi_history not found'}), 404
    return jsonify({'label': d.get('label', '소비자물가상승률(%)'),
                    'points': _cpi_graph_points(d),
                    'annual': d.get('annual', {}), 'monthly': d.get('monthly', {})})


@app.route('/download/cpi_history')
def download_cpi_history():
    """'과거 물가상승률 다운로드' → 누적된 과거 물가상승률 엑셀 생성(연도별 + 올해 월별)."""
    import openpyxl
    d = _cpi_history_load()
    if not d:
        return jsonify({'error': 'cpi_history not found'}), 404
    annual = d.get('annual', {})
    monthly = d.get('monthly', {})
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '소비자물가상승률'
    hdr = ['']
    vals = [d.get('label', '소비자물가상승률(%)')]
    for y in sorted(annual):
        hdr.append(y)
        vals.append(annual[y])
    for k in sorted(monthly):
        y, m = k.split('-')
        hdr.append('%s년 %d월' % (y, int(m)))
        vals.append(monthly[k])
    ws.append(hdr)
    ws.append(vals)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name='과거 물가상승률.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/pension_export', methods=['POST'])
def pension_export():
    """당월/전체 금리표(화면 형식) 엑셀 생성. 월별 1시트 + 금리연동형/기타 섹션 포함.
       payload: {filename, months:[{month:'YYYY-MM', rows:[{sector,org,fam,db,dc,def}], special:{rateLinked,period}}]}"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    payload = request.get_json(force=True, silent=True) or {}
    months = payload.get('months', [])
    if not months:
        return jsonify({'error': 'no data'}), 400
    MO = _PENSION_MONTHS                      # [3,6,12,18,24,30,36,48,60]
    MLABEL = {3: '3개월', 6: '6개월', 12: '1년', 18: '18개월', 24: '2년',
              30: '30개월', 36: '3년', 48: '4년', 60: '5년'}
    NC = 3 + len(MO) * 2 + 1                  # 22열 (업권·기관·상품 + DB9 + DC9 + 디폴트)

    thin = Side(style='thin', color='D0D5DD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='F1F3F5')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center')

    def num(v):
        try:
            if v is None or v == '' or v == '-':
                return None
            return round(float(v), 2)
        except (TypeError, ValueError):
            return None

    def style_row(ws, r, ncol, leftcols):
        for cc in range(1, ncol + 1):
            c = ws.cell(row=r, column=cc)
            c.border = border
            c.font = Font(size=9)
            c.alignment = left if cc in leftcols else center

    def section_header(ws, r, headers):
        for j, h in enumerate(headers):
            c = ws.cell(row=r, column=1 + j, value=h)
            c.fill = hdr_fill
            c.font = Font(bold=True, size=9, color='374151')
            c.alignment = center
            c.border = border

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for entry in months:
        m = str(entry.get('month', ''))
        rows = entry.get('rows', []) or []
        special = entry.get('special', {}) or {}
        try:
            y, mo = m.split('-')
            label = '%s년 %d월' % (y, int(mo))
        except Exception:
            label = m or '금리현황'
        ws = wb.create_sheet(title=(label or '금리현황')[:31])

        # 제목
        tc = ws.cell(row=1, column=1, value='■ %s 퇴직연금 원리금보장상품 금리 현황' % label)
        tc.font = Font(bold=True, size=13)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NC)

        # 헤더(3~4행): 값 기입 → 스타일 → 병합
        hr1, hr2 = 3, 4
        ws.cell(row=hr1, column=1, value='업권')
        ws.cell(row=hr1, column=2, value='상품제공기관')
        ws.cell(row=hr1, column=3, value='상품구분')
        ws.cell(row=hr1, column=4, value='DB')
        ws.cell(row=hr1, column=4 + len(MO), value='DC / IRP')
        ws.cell(row=hr1, column=NC, value='디폴트\n옵션용 3년')
        for i, mm in enumerate(MO):
            ws.cell(row=hr2, column=4 + i, value=MLABEL[mm])
            ws.cell(row=hr2, column=4 + len(MO) + i, value=MLABEL[mm])
        for rr in (hr1, hr2):
            for cc in range(1, NC + 1):
                c = ws.cell(row=rr, column=cc)
                c.fill = hdr_fill
                c.font = Font(bold=True, size=9, color='374151')
                c.alignment = center
                c.border = border
        ws.merge_cells(start_row=hr1, start_column=1, end_row=hr2, end_column=1)
        ws.merge_cells(start_row=hr1, start_column=2, end_row=hr2, end_column=2)
        ws.merge_cells(start_row=hr1, start_column=3, end_row=hr2, end_column=3)
        ws.merge_cells(start_row=hr1, start_column=4, end_row=hr1, end_column=3 + len(MO))
        ws.merge_cells(start_row=hr1, start_column=4 + len(MO), end_row=hr1, end_column=3 + len(MO) * 2)
        ws.merge_cells(start_row=hr1, start_column=NC, end_row=hr2, end_column=NC)

        # 데이터
        dr = hr2 + 1
        for r in rows:
            db = r.get('db', {}) or {}
            dc = r.get('dc', {}) or {}
            ws.cell(row=dr, column=1, value=r.get('sector', ''))
            ws.cell(row=dr, column=2, value=r.get('org', ''))
            ws.cell(row=dr, column=3, value=r.get('fam', ''))
            for i, mm in enumerate(MO):
                ws.cell(row=dr, column=4 + i, value=num(db.get(str(mm), db.get(mm))))
                ws.cell(row=dr, column=4 + len(MO) + i, value=num(dc.get(str(mm), dc.get(mm))))
            ws.cell(row=dr, column=NC, value=num(r.get('def')))
            style_row(ws, dr, NC, (2, 3))
            for cc in range(4, NC + 1):
                ws.cell(row=dr, column=cc).number_format = '0.00'
            dr += 1

        # 금리연동형
        rl = special.get('rateLinked', []) or []
        if rl:
            dr += 1
            ws.cell(row=dr, column=1, value='· 금리연동형').font = Font(bold=True, size=10)
            dr += 1
            section_header(ws, dr, ['업권', '상품제공기관', '상품구분', '금리'])
            dr += 1
            for r in rl:
                ws.cell(row=dr, column=1, value=r.get('sector', ''))
                ws.cell(row=dr, column=2, value=r.get('org', ''))
                ws.cell(row=dr, column=3, value=r.get('fam', ''))
                ws.cell(row=dr, column=4, value=num(r.get('rate')))
                style_row(ws, dr, 4, (2, 3))
                ws.cell(row=dr, column=4).number_format = '0.00'
                dr += 1

        # 기타(만기지정식·일단위지정)
        pd = special.get('period', []) or []
        if pd:
            dr += 1
            ws.cell(row=dr, column=1, value='· 기타 (만기지정식·일단위지정)').font = Font(bold=True, size=10)
            dr += 1
            section_header(ws, dr, ['업권', '상품제공기관', '상품구분', '만기(개월)', 'DB', 'DC', 'IRP'])
            dr += 1
            for r in pd:
                ws.cell(row=dr, column=1, value=r.get('sector', ''))
                ws.cell(row=dr, column=2, value=r.get('org', ''))
                ws.cell(row=dr, column=3, value=r.get('fam', ''))
                ws.cell(row=dr, column=4, value=r.get('mat', ''))
                for k, key in enumerate(('db', 'dc', 'irp')):
                    ws.cell(row=dr, column=5 + k, value=num(r.get(key)))
                style_row(ws, dr, 7, (2, 3))
                for cc in range(5, 8):
                    ws.cell(row=dr, column=cc).number_format = '0.00'
                dr += 1

        # 열 너비
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 22
        for i in range(len(MO) * 2):
            ws.column_dimensions[get_column_letter(4 + i)].width = 8
        ws.column_dimensions[get_column_letter(NC)].width = 10

    if not wb.sheetnames:
        wb.create_sheet(title='금리현황')
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = payload.get('filename') or '퇴직연금 금리현황.xlsx'
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 원리금보장 금리현황 레포트: 양식(pension_report_template.xlsx)에 현재 화면 금리 채워 반환 ──
_PENSION_MONTHS = [3, 6, 12, 18, 24, 30, 36, 48, 60]
_PENSION_ALIAS = {  # 레포트 표기 -> 데이터 표기(예외)
    '기업은행': '중소기업은행', '산업은행': '한국산업은행', 'SBI저축은행': '에스비아이저축은행',
    'NH저축은행': '엔에이치저축은행', 'DB저축은행': '디비저축은행', 'JT친애저축은행': '제이티친애저축은행',
    'BNK투자증권': '비엔케이투자증권', '신한생명': '신한라이프',
    'SC은행': '한국스탠다드차타드은행',   # 레포트 SC은행 = 데이터 한국스탠다드차타드은행(SC제일은행 정기예금)
}
# 원리금 데이터 기관명 → 신용등급관리 기관명(모바일 상세 팝업 신용등급 매칭용).
#   이름이 달라 자동 매칭이 안 되는 기관은 지정한 신용등급 기관의 등급을 끌어온다.
_PEN_RATING_ALIAS = {
    '한국스탠다드차타드은행': 'SC제일은행',
    'IBK투자증권': '아이비케이투자증권',
}
_PENSION_SECMAP = {'증권': '증권', '은행': '은행', '생보': '생명보험', '손보': '손해보험', '저축은행': '저축은행'}
_ROMAN = [('Ⅲ', '3'), ('Ⅱ', '2'), ('Ⅰ', '1'), ('Ⅳ', '4'), ('Ⅴ', '5'),
          ('ⅲ', '3'), ('ⅱ', '2'), ('ⅰ', '1'), ('III', '3'), ('II', '2'), ('IV', '4'), ('V', '5'), ('I', '1')]


def _pen_core(s):
    s = str(s).lower()
    s = re.sub(r'주식회사|㈜|\(주\)|\s', '', s)
    s = re.sub(r'(생명보험|손해보험|화재보험|연금보험)$', '', s)
    s = re.sub(r'(화재|생명|손보|손해|연금)$', '', s)
    return s


def _pen_sec_token(sector):
    """업권 구분 토큰. 손보/생명은 같은 브랜드(삼성·한화·DB·하나 등)가 양쪽에
    모두 있어 _pen_core만으로는 'db'처럼 키가 충돌한다. 손보/생명 등급이 서로
    덮어쓰이지 않도록 정규화 키 앞에 업권 토큰을 붙여 구분한다.
    pension.html penSecTok()와 동일 규칙이어야 한다(변경 시 양쪽 같이 수정)."""
    if sector == '손해보험':
        return '손보'
    if sector == '생명보험':
        return '생명'
    return ''


def _pen_rating_key(sector, name):
    """신용등급 맵의 조회 키 = 업권토큰 + 정규화기관명."""
    return _pen_sec_token(sector) + _pen_core(name)


def _pen_inst_key(name, is_report=False):
    if is_report and name in _PENSION_ALIAS:
        name = _PENSION_ALIAS[name]
    return _pen_core(name)


def _pen_secnorm(s):
    return _PENSION_SECMAP.get(str(s).replace('\n', '').strip(), str(s))


def _pen_roman(s):
    s = str(s)
    for k, v in _ROMAN:
        s = s.replace(k, v)
    return s


def _pen_iyul(fam):
    f = _pen_roman(fam)
    bonus = '보너스' in f
    base = re.sub(r'/.*$', '', f)
    m = re.search(r'이율보증형(보험)?\s*([123])', base)
    if m:
        return (bonus, m.group(2))
    if '이율보증형' in base:
        return (bonus, '1')
    return (None, None)


def _pen_match_product(rp, g):
    fam = g['fam']
    text = fam + ' ' + g['names']
    rp = rp.strip()
    if rp == '정기예금':
        return '정기예금' in text
    if rp == 'ELB':
        return ('ELB' in text) and (('ELB' in fam) or ('DLB' not in fam))
    if rp == 'DLB':
        return 'DLB' in text
    if rp == 'ELB/DLB':
        return ('ELB' in text) or ('DLB' in text)
    if rp == '발행어음':
        return '발행어음' in text
    if rp == 'RP':
        return (fam == 'RP') or ('RP' in fam) or ('환매조건부' in text)
    if '보너스이율보증형' in rp:
        b, _ = _pen_iyul(fam)
        return b is True
    m = re.search(r'이율보증형보험\s*([123])?', rp)
    if m:
        want = m.group(1) or '1'
        b, num = _pen_iyul(fam)
        return (b is False) and (num == want)
    return bool(rp) and rp in text


@app.route('/api/pension_ratings')
def api_pension_ratings():
    """상품제공기관 정규화키 -> 최종 신용등급 맵.

    모바일 금리표 상세 팝업에서 기관의 신용등급을 표시할 때 사용한다.
    키는 _pen_core()로 정규화한 기관명(주식회사/보험류 접미 제거)이며,
    pension.html의 penCore()와 동일 규칙이어야 한다(변경 시 양쪽 같이 수정).
    """
    data = build_response_data(load_institutions(), load_ratings())
    out = {}
    for category, rows in data.items():
        for row in rows:
            if row.get('final'):
                out[_pen_rating_key(category, row['name'])] = row['final']
    # 별칭: 원리금 데이터 기관명이 신용등급 기관명과 달라 매칭 안 되는 경우
    #       지정한 신용등급 기관의 등급을 그 이름 키로도 넣어준다(SC·IBK 등).
    #       별칭 대상은 은행·증권(업권토큰 없음)이라 정규화 키만으로 매칭한다.
    for pen_name, cred_name in _PEN_RATING_ALIAS.items():
        ck = _pen_core(cred_name)
        if ck in out:
            out[_pen_core(pen_name)] = out[ck]
    return jsonify(out)


def _normalize_pen_rows(rows):
    """클라이언트 그룹 rows(curRows) → 서버 계산용 G(db/dc 키 int화)."""
    G = []
    for r in rows:
        db = {int(k): v for k, v in (r.get('db') or {}).items() if v is not None}
        dc = {int(k): v for k, v in (r.get('dc') or {}).items() if v is not None}
        G.append({'sector': r.get('sector'), 'org': r.get('org') or '', 'fam': r.get('fam') or '',
                  'names': ' '.join(r.get('names') or []), 'db': db, 'dc': dc, 'def': r.get('def')})
    return G


def _build_pension_report_wb(rows, month):
    """rows(curRows)를 레포트 양식에 채운 워크북 반환 → (wb, ws, matched).
    금리현황 다운로드와 과거 금리 추이 이력계산이 '동일한 채움 결과'를 쓰도록 공용화."""
    import openpyxl
    from openpyxl.cell.cell import MergedCell
    G = _normalize_pen_rows(rows)

    def find_rp(ikey):
        c = [g for g in G if _pen_inst_key(g['org']) == ikey
             and (g['fam'] == 'RP' or 'RP' in g['fam'] or '환매조건부' in (g['fam'] + g['names']) or '발행어음' in (g['fam'] + g['names']))]
        c.sort(key=lambda x: len(x['db']) + len(x['dc']), reverse=True)
        return c[0] if c else None

    wb = openpyxl.load_workbook(os.path.join(BASE_DIR, 'pension_report_template.xlsx'))
    ws = wb.worksheets[0]
    # 디폴트옵션(X=24열) 섹션 병합(증권 X7:35·생보 X53:70·손보 X71:79) 해제
    #  → 보험사별 첫 '이율보증형보험' 행에 각자의 디폴트옵션 금리를 개별 기록(병합이면 스킵되던 문제 해결)
    for _rng in ('X7:X35', 'X53:X70', 'X71:X79'):
        try:
            ws.unmerge_cells(_rng)
        except (KeyError, ValueError):
            pass

    def setcell(rr, cc, v):
        cell = ws.cell(rr, cc)
        if isinstance(cell, MergedCell):   # 병합 비앵커 셀은 건너뜀(디폴트 X 섹션병합 등)
            return
        cell.value = v

    if month and re.match(r'\d{4}-\d{2}', month):
        y, mm = month.split('-')
        setcell(1, 2, f"          퇴직연금 원리금지급형상품 공시금리 현황 [{y}년 {int(mm):02d}월]                ")

    DB_C0, DC_C0, X_C, RP_C = 6, 15, 24, {'db': 26, 'dc': 27, 'irp': 28}
    cur_sec = cur_inst = None
    matched = 0
    for r in range(7, ws.max_row + 1):   # 양식 행 추가/변경에도 대응(상품구분 없는 행은 스킵)
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        e = ws.cell(r, 5).value
        if b:
            cur_sec = b
        if c:
            cur_inst = c
        if not e:
            continue
        ikey = _pen_inst_key(cur_inst, is_report=True)
        dsec = _pen_secnorm(cur_sec)
        cand = [g for g in G if _pen_secnorm(g['sector']) == dsec and _pen_inst_key(g['org']) == ikey]
        if not cand:
            cand = [g for g in G if _pen_inst_key(g['org']) == ikey]
        if not cand:
            continue
        pm = [g for g in cand if _pen_match_product(str(e), g)]
        if not pm:
            continue
        pm.sort(key=lambda x: len(x['db']) + len(x['dc']) + (1 if x['def'] else 0), reverse=True)
        mdb, mdc, mdef = {}, {}, None
        for g in pm:
            for m in _PENSION_MONTHS:
                if m in g['db'] and m not in mdb:
                    mdb[m] = g['db'][m]
                if m in g['dc'] and m not in mdc:
                    mdc[m] = g['dc'][m]
            if mdef is None and g['def'] is not None:
                mdef = g['def']
        matched += 1
        for i, m in enumerate(_PENSION_MONTHS):
            if m in mdb:
                setcell(r, DB_C0 + i, round(float(mdb[m]), 3))
            if m in mdc:
                setcell(r, DC_C0 + i, round(float(mdc[m]), 3))
        if mdef is not None:
            setcell(r, X_C, round(float(mdef), 3))
        if c and dsec == '증권':   # RP금리(1년): 기관 첫 행에 기록
            rp = find_rp(ikey)
            if rp:
                if 12 in rp['db']:
                    setcell(r, RP_C['db'], round(float(rp['db'][12]), 3))
                if 12 in rp['dc']:
                    setcell(r, RP_C['dc'], round(float(rp['dc'][12]), 3))
                    setcell(r, RP_C['irp'], round(float(rp['dc'][12]), 3))
    return wb, ws, matched


@app.route('/api/pension_report', methods=['POST'])
def pension_report():
    """현재 화면(기준월)의 금리를 레포트 양식에 채워 xlsx로 반환."""
    import io as _io
    payload = request.get_json(force=True) or {}
    month = payload.get('month', '')
    rows = payload.get('rows', [])
    wb, ws, matched = _build_pension_report_wb(rows, month)
    bio = _io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"퇴직연금 금리정보 현황_{month}.xlsx" if month else "퇴직연금 금리정보 현황.xlsx"
    logger.info('금리현황 레포트 생성: month=%s, 데이터 %d행, 매칭 %d행', month, len(rows), matched)
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 과거 금리 추이: 매월 업권별 DB1년 평균 자동 append ──
#   리포트 H열(=DB 1년)의 업권별 기관행 범위 평균(사용자 지정 2026-08).
#   ※ 이 범위는 pension_report_template의 고정 행 배치에 대응(증권7~, 은행36~, ...).
_RH_SECTOR_ROWS = [('증권', 7, 34), ('은행', 36, 52), ('생보', 53, 70),
                   ('손보', 71, 79), ('저축', 80, 101)]
_RH_DB_1Y_COL = 8   # 리포트 H열 = DB 1년


def _avg_report_col(ws, col, r0, r1):
    vals = [float(v) for r in range(r0, r1 + 1)
            for v in [ws.cell(r, col).value]
            if isinstance(v, (int, float)) and not isinstance(v, bool)]
    return round(sum(vals) / len(vals), 6) if vals else None


@app.route('/api/rate_history_append', methods=['POST'])
def rate_history_append():
    """현재 리포트(rows)에서 업권별 DB1년 평균을 계산해 과거 금리 추이(Sheet2)에 그달 한 줄 추가.
    행: [DATE(그달1일), 증권, 은행, 생보, 손보, 저축, 원리금보장평균(5개 평균), 기준금리]. 같은 DATE는 교체."""
    import openpyxl
    payload = request.get_json(force=True) or {}
    month = str(payload.get('month', ''))
    rows = payload.get('rows', [])
    if not re.match(r'^\d{4}-\d{2}$', month):
        return jsonify({'error': '기준월(YYYY-MM)이 필요합니다'}), 400
    try:
        base_rate = float(payload.get('base_rate'))
    except (TypeError, ValueError):
        return jsonify({'error': '기준금리(숫자)를 입력해주세요'}), 400

    wb, ws, matched = _build_pension_report_wb(rows, month)
    secavg = {name: _avg_report_col(ws, _RH_DB_1Y_COL, r0, r1) for name, r0, r1 in _RH_SECTOR_ROWS}
    # 원리금보장 평균 = 증권·은행·생보·손보 4개 평균(저축은행 제외 — 과거 데이터 산식과 동일)
    four = [secavg[n] for n in ('증권', '은행', '생보', '손보')]
    present = [v for v in four if v is not None]
    overall = round(sum(present) / len(present), 6) if present else None
    date_str = f"{month}-01"
    new_row = [date_str, secavg['증권'], secavg['은행'], secavg['생보'],
               secavg['손보'], secavg['저축'], overall, base_rate]

    def dstr(v):
        return str(v)[:10] if v is not None else ''

    # json(Sheet2) 갱신 — 같은 DATE 있으면 교체, 없으면 추가 후 날짜순 정렬
    with open(_RATE_HISTORY_JSON, encoding='utf-8') as f:
        hist = json.load(f)
    rj = hist['sheets']['Sheet2']['rows']
    ix = next((i for i, rr in enumerate(rj) if dstr(rr[0]) == date_str), None)
    if ix is None:
        rj.append(new_row)
    else:
        rj[ix] = new_row
    rj.sort(key=lambda rr: dstr(rr[0]))
    with open(_RATE_HISTORY_JSON, 'w', encoding='utf-8') as f:
        json.dump(hist, f, ensure_ascii=False)

    # xlsx(Sheet2) 갱신 — 같은 DATE 있으면 그 행 교체, 없으면 맨 아래 추가(날짜는 datetime으로 기록)
    import datetime as _dt
    y, mm = month.split('-')
    dt_val = _dt.datetime(int(y), int(mm), 1)
    xb = openpyxl.load_workbook(_RATE_HISTORY_XLSX)
    xs = xb['Sheet2']
    target = None
    for r in range(5, xs.max_row + 1):
        if dstr(xs.cell(r, 1).value) == date_str:
            target = r
            break
    if target is None:
        target = xs.max_row + 1
    xs.cell(target, 1, dt_val)
    for ci, val in enumerate(new_row[1:], start=2):
        xs.cell(target, ci, val)
    # 서식을 바로 위 데이터 행과 동일하게 맞춤 → 날짜(시간 없이)·금리 소수점 2자리로 7월 행처럼 표시
    ref = target - 1
    if ref >= 5:
        for ci in range(1, 9):
            xs.cell(target, ci).number_format = xs.cell(ref, ci).number_format
    xb.save(_RATE_HISTORY_XLSX)

    logger.info('과거금리추이 append %s: 증권=%s 은행=%s 생보=%s 손보=%s 저축=%s 평균=%s 기준=%s (매칭 %d)',
                date_str, secavg['증권'], secavg['은행'], secavg['생보'], secavg['손보'],
                secavg['저축'], overall, base_rate, matched)
    return jsonify({'ok': True, 'date': date_str, 'sector': secavg, 'overall': overall,
                    'base_rate': base_rate, 'matched': matched})


@app.route('/simple2')
def simple2_preview():
    # 모바일 ver2 미리보기(개발용) — 로그인 세션 공유로 /api/* 정상 동작
    return _record_visit(make_response(send_file(os.path.join(BASE_DIR, 'simple2.html'))))


@app.route('/')
def index():
    # 간편조회(v2, 모바일 신규 디자인) — 로그인 시 선택. 기존 v1은 아래 그대로 유지.
    if session.get('view') == 'simple':
        return _record_visit(make_response(send_file(os.path.join(BASE_DIR, 'simple.html'))))
    institutions = load_institutions()
    ratings = load_ratings()
    data = build_response_data(institutions, ratings)
    meta = ratings.get('_meta', {})

    # 변경 알람 요약 + 변경 감지된 기관 목록(이름·카테고리)
    changed_institutions = [
        {'name': row['name'], 'category': category}
        for category, rows in data.items()
        for row in rows if row['any_changed']
    ]
    total_changed = len(changed_institutions)

    # 변경이력 (최신순)
    history = sorted(
        load_history(),
        key=lambda h: (h.get('detected_at', ''), h.get('eval_date', '')),
        reverse=True,
    )

    html = render_template(
        'index.html',
        data=data,
        last_updated=meta.get('updated', '-'),
        scrape_summary=meta.get('scrape_summary', ''),
        total_changed=total_changed,
        changed_institutions=changed_institutions,
        history=history,
        rating_scale=RATING_SCALE,
        agencies=AGENCIES,
        agency_labels=AGENCY_LABELS,
        role=session.get('role', ''),   # consulting=전체, rm=조회·다운로드만(화면 제어용)
        rm_user=RM_USER,                 # RM 비밀번호 변경 UI 표시/대상용(미설정 시 버튼 숨김)
        rm_days_left=_rm_days_left(),    # RM 비밀번호 만료까지 남은 일수(표시용)
    )
    # 메인 페이지 로드 = 1 접속 → 일자별 방문자 집계(웹/모바일 모두)
    return _record_visit(make_response(html))


@app.route('/switch_view')
def switch_view():
    """간편조회(v2) ↔ 웹버전조회(v1) 전환. 로그아웃 없이 토글."""
    session['view'] = 'web' if session.get('view') == 'simple' else 'simple'
    return redirect(url_for('index'))


@app.route('/api/me')
def api_me():
    """현재 로그인 사용자/소속(role) + 조회방식(view). iframe/화면 제어용."""
    return jsonify({'user': session.get('user', ''), 'role': session.get('role', ''),
                    'view': session.get('view', 'web')})


# ── 방문자 통계 (연금컨설팅팀 전용 관리자 화면) ─────────────────────────
#   _ADMIN_ONLY_ENDPOINTS 에 등록됨 → RM 계정은 서버측에서 차단.
@app.route('/admin/visitors')
def admin_visitors():
    return send_file(os.path.join(BASE_DIR, 'admin_visitors.html'))


@app.route('/api/visit_stats')
def api_visit_stats():
    return jsonify({'series': _visit_series()})


@app.route('/download/visit_stats.xlsx')
def download_visit_stats():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = '방문자통계'
    headers = ['일자', '웹 접속수', '모바일 접속수', '총 접속수',
               '웹 순방문', '모바일 순방문', '총 순방문']
    ws.append(headers)
    navy = PatternFill('solid', fgColor='16335B')
    thin = Side(style='thin', color='D5D9E0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = navy
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    for r in _visit_series():
        ws.append([r['date'], r['web_total'], r['mob_total'], r['total'],
                   r['web_uniq'], r['mob_uniq'], r['uniq']])
    for row in ws.iter_rows(min_row=2):
        for i, c in enumerate(row):
            c.border = border
            c.alignment = Alignment(horizontal='left' if i == 0 else 'right')
    ws.column_dimensions['A'].width = 13
    for col in ('B', 'C', 'D', 'E', 'F', 'G'):
        ws.column_dimensions[col].width = 13
    ws.freeze_panes = 'A2'
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = '방문자통계_%s.xlsx' % datetime.now().strftime('%Y%m%d')
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 서버 배포 (웹 「서버배포」 버튼 → VM이 코드 pull+재시작) ─────────────
#   credit-deploy.service(oneshot)를 비동기(--no-block)로 트리거만 한다.
#   실제 pull/재시작은 auto_deploy.sh가 수행(코드만, 데이터 미변경).
#   진행상태는 auto_deploy.sh가 deploy_status.json에 단계별로 기록 → 프런트가 /admin/deploy_status 폴링.
#   최초 1회 VM 설치(deploy/install_deploy.sh) 후 동작. sudoers로 이 명령만 무비번 허용.
DEPLOY_STATUS_FILE = os.path.join(BASE_DIR, 'deploy_status.json')


def _write_deploy_status(state, message=''):
    try:
        with open(DEPLOY_STATUS_FILE, 'w', encoding='utf-8') as f:
            json.dump({'state': state, 'message': message,
                       'ts': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
                      f, ensure_ascii=False)
    except Exception:
        logger.exception('deploy_status 기록 실패')


def _read_deploy_status():
    try:
        with open(DEPLOY_STATUS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {'state': 'idle', 'message': '', 'ts': ''}


@app.route('/admin/deploy', methods=['POST'])
def admin_deploy():
    # systemd 서비스 환경은 PATH가 제한적이라 sudo/systemctl을 절대경로로 호출
    sudo = shutil.which('sudo') or '/usr/bin/sudo'
    systemctl = shutil.which('systemctl') or '/usr/bin/systemctl'
    # 이전 배포의 종료상태가 남아 프런트가 즉시 '완료'로 오인하지 않도록 먼저 비종료 상태로 초기화
    _write_deploy_status('queued', '배포 요청됨 — 곧 시작합니다')
    try:
        subprocess.Popen([sudo, systemctl, '--no-block', 'start', 'credit-deploy.service'])
    except Exception as e:
        logger.exception('배포 트리거 실패')
        _write_deploy_status('error', '배포 시작 실패(서버 설정 확인)')
        return jsonify({'success': False,
                        'message': '배포 시작 실패(서버 설정 확인): %s' % e}), 500
    return jsonify({'success': True,
                    'message': '배포를 시작했습니다.'})


@app.route('/admin/deploy_status')
def admin_deploy_status():
    """배포 진행 상태 조회(프런트 폴링용). state: idle|queued|running|restarting|success|no-change|merge-failed|restart-failed|error"""
    return jsonify(_read_deploy_status())


@app.route('/api/export.xlsx')
def api_export_xlsx():
    """조회된 전체 기관 정보를 첨부 양식(export_template.xlsx)에 채워 다운로드.
    컬럼: 업권 | 기관명 | 나이스신용평가 | 한국기업평가 | 한국신용평가 | 최종적용신용등급 | 최종등급 근거
    """
    import openpyxl
    from copy import copy

    institutions = load_institutions()
    ratings = load_ratings()
    data = build_response_data(institutions, ratings)

    wb = openpyxl.load_workbook(EXPORT_TEMPLATE)
    ws = wb.active

    # 1행(헤더)은 유지, 2행 이후를 데이터로 채움.
    # 템플릿 2행에 들어있던 예시 스타일을 복제해 각 데이터 행에 적용.
    sample_styles = [copy(ws.cell(row=2, column=c)._style) for c in range(1, 8)]
    # 기존 데이터 영역(2행~) 비우기
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    r = 2
    for category, rows in data.items():
        for row in rows:
            ags = row['agencies']
            values = [
                category,
                row['name'],
                ags['nice']['rating'],
                ags['kr']['rating'],
                ags['kis']['rating'],
                row['final'],
                row['basis_agency'],
            ]
            for c, v in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell._style = copy(sample_styles[c - 1])
            r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"신용등급_조회결과_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname,
    )


@app.route('/api/ratings')
def api_ratings():
    institutions = load_institutions()
    ratings = load_ratings()
    data = build_response_data(institutions, ratings)
    meta = ratings.get('_meta', {})
    return jsonify({'data': data, 'last_updated': meta.get('updated', '-')})


def _apply_scrape_results(scrape_result: dict, update_meta: bool = True,
                          alive_override: set | None = None) -> tuple[int, int]:
    """스크래핑 결과를 ratings.json에 반영. (success_count, changed_count) 반환.

    - update_meta: False면 전역 요약/타임스탬프(_meta)를 갱신하지 않음 (단일 기관 재조회용).
    - alive_override: 지정 시 이 평가사 집합을 'alive'로 간주 (단일 기관 재조회는 교차 판단이
      불가하므로, 전체 조회가 해당 기관을 처리하는 것과 동일하게 3사를 alive로 넘겨 사용).
    """
    ratings = load_ratings()
    overrides = load_overrides()
    success_count = 0
    changed_count = 0
    history_records = []  # 이번 실행에서 감지된 등급 변경 이력
    _now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # 평가사별 '이번 실행 전체 생존' 판단: 어느 한 기관에서라도 값을 반환했으면 정상.
    # 특정 평가사가 모든 기관에서 빈값이면 사이트/API 전면 장애로 보고 stale 값을 지우지 않는다.
    if alive_override is not None:
        agency_alive = {ag: (ag in alive_override) for ag in AGENCIES}
    else:
        agency_alive = {ag: False for ag in AGENCIES}
        for _n, _d in scrape_result.items():
            if _n == '_meta':
                continue
            for ag in AGENCIES:
                if _d.get(ag, ''):
                    agency_alive[ag] = True
    for name, new_data in scrape_result.items():
        if name == '_meta':
            continue
        prev_entry = ratings.get(name, {})
        entry = dict(prev_entry)
        # 신규 기관 첫 스크래핑 시 일부 기관 실패해도 키가 항상 존재하도록 보장
        for _ag in AGENCIES:
            entry.setdefault(_ag, '')
            entry.setdefault(f'{_ag}_eval_date', '')
            entry.setdefault(f'{_ag}_type', '')
            entry.setdefault(f'{_ag}_prev', '')
            entry.setdefault(f'{_ag}_changed', False)
            entry.setdefault(f'{_ag}_miss', 0)
        any_ag_updated = False
        for ag in AGENCIES:
            old_rating = prev_entry.get(ag, '')
            new_rating = new_data.get(ag, '')
            new_date   = new_data.get(f'{ag}_eval_date', '')
            if new_rating:
                # NICE에서 평가일 없는 값은 '검색결과 테이블 폴백'(불안정 출처)에서 나온 것.
                # 이런 값으로는 '변경'을 표시하지 않음 → 과거 오파싱값(예: KB국민은행 AA+)과
                # 비교해 생기는 허위 변경(오탐)을 방지. KR은 항상 날짜 있음, KIS는 영향 없음.
                unreliable_nice = (ag == 'nice' and not new_date)
                if old_rating and old_rating != new_rating and not unreliable_nice:
                    entry[f'{ag}_prev'] = old_rating
                    entry[f'{ag}_changed'] = True
                    changed_count += 1
                    history_records.append({
                        'name': name,
                        'agency': ag,
                        'agency_label': AGENCY_LABELS[ag],
                        'prev': old_rating,
                        'current': new_rating,
                        'direction': compare_ratings(old_rating, new_rating),
                        'type_label': get_rating_type_label(new_data.get(f'{ag}_type', '')),
                        'eval_date': new_date,        # 변경일(평가일)
                        'detected_at': _now,          # 감지일시
                    })
                entry[ag] = new_rating
                entry[f'{ag}_eval_date'] = new_data.get(f'{ag}_eval_date', prev_entry.get(f'{ag}_eval_date', ''))
                entry[f'{ag}_type']      = new_data.get(f'{ag}_type',      prev_entry.get(f'{ag}_type', ''))
                any_ag_updated = True
        # stale 잔존값 정리: 이 기관 조회가 성공(다른 평가사 값 확보)했고, 해당 평가사가
        # 이번 실행에서 다른 기관들엔 정상 응답했는데 이 기관에서만 빈값이면 → 과거 오파싱으로
        # 남은 값(예: KB손해보험 KIS 'AAA')을 제거한다. 평가사 전면 장애(agency_alive=False)나
        # 기관 전체 실패(any_ag_updated=False) 시에는 보존한다.
        #
        # 단, 빈값 1회로 바로 지우면 일시적 스크래핑 실패(NICE 타임아웃 등)에 멀쩡한 등급이
        # 날아간다(2026-07-21 삼성증권 AA+·NH투자증권 AA+·기업은행 AAA 사례).
        # → 연속 MISS_LIMIT회 빈값일 때만 삭제하고, 그 전까지는 값을 유지한다.
        if any_ag_updated:
            for ag in AGENCIES:
                if new_data.get(ag, ''):
                    entry[f'{ag}_miss'] = 0                     # 값이 오면 카운터 리셋
                elif agency_alive[ag] and entry.get(ag, ''):
                    misses = int(prev_entry.get(f'{ag}_miss', 0) or 0) + 1
                    if misses >= MISS_LIMIT:
                        entry[ag] = ''
                        entry[f'{ag}_eval_date'] = ''
                        entry[f'{ag}_type'] = ''
                        entry[f'{ag}_miss'] = 0
                        logger.info('[%s] %s 연속 %d회 빈값 → 등급 삭제', name, AGENCY_LABELS[ag], misses)
                    else:
                        entry[f'{ag}_miss'] = misses            # 값 보존, 다음 조회에서 재판단
                        logger.info('[%s] %s 빈값 %d/%d회 — 기존 등급 유지',
                                    name, AGENCY_LABELS[ag], misses, MISS_LIMIT)
        old_final = prev_entry.get('final', '')
        new_finals = [entry.get(ag, '') for ag in AGENCIES]
        new_final  = get_lowest_rating(new_finals)
        if old_final and new_final and old_final != new_final:
            entry['final_prev']    = old_final
            entry['final_changed'] = True
        if new_final:
            entry['final'] = new_final
        status = new_data.get('scrape_status', '')
        if any_ag_updated:
            entry['scrape_status'] = status
            entry['updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            success_count += 1
        else:
            import re as _re
            base = _re.sub(r'\(재조회실패\)', '', prev_entry.get('scrape_status', '')).strip()
            entry['scrape_status'] = (base or '등급없음') + '(재조회실패)'
        # 사용자 강제 정정(override) 적용 → 재조회 결과가 정정값을 덮어쓰지 못하게 함
        if overrides.get(name):
            _apply_overrides(name, entry, overrides)
            entry['final'] = get_lowest_rating([entry.get(ag, '') for ag in AGENCIES])
        ratings[name] = entry
    if update_meta:
        ratings['_meta'] = {
            'updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'scrape_summary': f'조회 성공 {success_count}건 / 등급 변경 감지 {changed_count}건',
        }
    save_ratings(ratings)
    append_history(history_records)
    return success_count, changed_count


@app.route('/api/refresh', methods=['POST'])
def api_refresh():
    global _refresh_state
    with _refresh_lock:
        if _refresh_state['running']:
            elapsed = time.time() - (_refresh_state.get('started_at') or 0)
            if elapsed < STALE_REFRESH_SEC:
                return jsonify({'success': False, 'message': '이미 조회 중입니다'}), 409
            # 여기 도달 = 시작한 지 오래됐는데 끝나지 않음 → 죽은 조회로 보고 새로 시작
            logger.warning('이전 조회가 %.0f분째 진행 중 — 멈춘 것으로 보고 새 조회를 시작합니다',
                           elapsed / 60)

    institutions = load_institutions()
    total = sum(len(items) for items in institutions.values())

    with _refresh_lock:
        _refresh_state = {
            'running': True, 'total': total, 'completed': 0,
            'results': {}, 'summary': '', 'updated_at': '', 'error': '',
            'started_at': time.time(),
        }

    def _run():
        global _refresh_state
        try:
            from scraper import scrape_all_ratings
            all_results = {}

            def on_progress(name, data):
                with _refresh_lock:
                    _refresh_state['results'][name] = data
                    _refresh_state['completed'] += 1
                all_results[name] = data

            scrape_all_ratings(institutions, progress_callback=on_progress)
            success_count, changed_count = _apply_scrape_results(all_results)
            summary = f'조회 성공 {success_count}건 / 등급 변경 감지 {changed_count}건'
            with _refresh_lock:
                _refresh_state['running']    = False
                _refresh_state['summary']    = summary
                _refresh_state['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        except Exception as e:
            logger.exception('Refresh failed')
            with _refresh_lock:
                _refresh_state['running'] = False
                _refresh_state['error']   = str(e)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({'success': True, 'total': total})


@app.route('/api/refresh_status')
def api_refresh_status():
    with _refresh_lock:
        return jsonify(dict(_refresh_state))


@app.route('/api/refresh_one/<path:name>', methods=['POST'])
def api_refresh_one(name):
    """기관 1개만 즉시 재조회 (전체 조회 없이). 전역 요약 배너는 갱신하지 않는다."""
    institutions = load_institutions()
    category = next(
        (cat for cat, items in institutions.items()
         if any(i['name'] == name for i in items)),
        None,
    )
    if category is None:
        return jsonify({'success': False, 'message': '기관을 찾을 수 없습니다'}), 404

    from scraper import _scrape_one, SAVING_BANK_CATEGORIES
    is_ins = category in INSURANCE_CATEGORIES
    is_sav = category in SAVING_BANK_CATEGORIES
    try:
        data = _scrape_one(name, is_ins, is_sav)
    except Exception as e:
        logger.exception('단일 재조회 실패 [%s]', name)
        return jsonify({'success': False, 'message': str(e)}), 500

    # 단일 기관이라 평가사 교차 생존 판단이 불가 → 전체 조회와 동일하게 3사를 alive로 취급.
    # 전역 _meta(마지막 업데이트/요약)는 건드리지 않음.
    _apply_scrape_results({name: data}, update_meta=False, alive_override=set(AGENCIES))

    ratings = load_ratings()
    row = build_row(name, category, ratings.get(name, {}), load_overrides())
    return jsonify({'success': True, 'row': row,
                    'scrape_status': ratings.get(name, {}).get('scrape_status', '')})


@app.route('/api/ratings/<path:name>', methods=['PUT'])
def api_update_rating(name):
    data = request.get_json()
    ratings = load_ratings()
    entry = ratings.get(name, {})

    for ag in AGENCIES:
        old_rating = entry.get(ag, '')
        new_rating = data.get(ag, '').strip()
        new_date   = data.get(f'{ag}_eval_date', '').strip()
        new_type   = data.get(f'{ag}_type', '').strip()

        if old_rating and new_rating and old_rating != new_rating:
            entry[f'{ag}_prev']    = old_rating
            entry[f'{ag}_changed'] = True
        entry[ag] = new_rating
        if new_date:
            entry[f'{ag}_eval_date'] = new_date
        if new_type:
            entry[f'{ag}_type'] = new_type

    # 최종 등급 갱신
    old_final = entry.get('final', '')
    new_final = get_lowest_rating([entry.get(ag, '') for ag in AGENCIES])
    if old_final and new_final and old_final != new_final:
        entry['final_prev']    = old_final
        entry['final_changed'] = True
    entry['final']   = new_final
    entry['updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    entry['scrape_status'] = '수동입력'

    if '_meta' not in ratings:
        ratings['_meta'] = {}
    ratings['_meta']['updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    ratings[name] = entry
    save_ratings(ratings)
    return jsonify({'success': True, 'final': new_final})


@app.route('/api/acknowledge/<path:name>', methods=['POST'])
def api_acknowledge(name):
    """변경 알람 확인 처리 (changed 플래그 초기화)"""
    ratings = load_ratings()
    if name in ratings:
        for ag in AGENCIES:
            ratings[name][f'{ag}_changed'] = False
        ratings[name]['final_changed'] = False
        save_ratings(ratings)
    return jsonify({'success': True})


@app.route('/api/acknowledge_all', methods=['POST'])
def api_acknowledge_all():
    ratings = load_ratings()
    for name, entry in ratings.items():
        if name == '_meta':
            continue
        for ag in AGENCIES:
            entry[f'{ag}_changed'] = False
        entry['final_changed'] = False
    save_ratings(ratings)
    return jsonify({'success': True})


@app.route('/api/institutions', methods=['POST'])
def api_add_institution():
    data = request.get_json()
    category = data.get('category', '').strip()
    name = data.get('name', '').strip()
    if not category or not name:
        return jsonify({'success': False, 'message': '카테고리와 기관명을 입력하세요'}), 400
    institutions = load_institutions()
    if category not in institutions:
        return jsonify({'success': False, 'message': '올바르지 않은 카테고리'}), 400
    if name in [i['name'] for i in institutions[category]]:
        return jsonify({'success': False, 'message': '이미 등록된 기관'}), 400
    institutions[category].append({'name': name})
    with open(INSTITUTIONS_FILE, 'w', encoding='utf-8') as f:
        json.dump(institutions, f, ensure_ascii=False, indent=2)
    return jsonify({'success': True})


@app.route('/api/institutions/<path:name>', methods=['DELETE'])
def api_delete_institution(name):
    institutions = load_institutions()
    found = False
    for cat in institutions:
        before = len(institutions[cat])
        institutions[cat] = [i for i in institutions[cat] if i['name'] != name]
        if len(institutions[cat]) < before:
            found = True
    if not found:
        return jsonify({'success': False, 'message': '기관 없음'}), 404
    with open(INSTITUTIONS_FILE, 'w', encoding='utf-8') as f:
        json.dump(institutions, f, ensure_ascii=False, indent=2)
    ratings = load_ratings()
    ratings.pop(name, None)
    save_ratings(ratings)
    return jsonify({'success': True})


# ── Scheduler ─────────────────────────────────────────────────────────

def scheduled_job():
    logger.info('스케줄 실행: 신용등급 자동 조회 시작')
    with app.app_context():
        try:
            import requests as req
            req.post(f'http://localhost:{PORT}/api/refresh', timeout=300)
        except Exception:
            logger.exception('스케줄 실행 오류')
        try:
            fetch_bond_rates()   # 시장금리 4종(ECOS) 갱신
        except Exception:
            logger.exception('시장금리(ECOS) 갱신 오류')


if __name__ == '__main__':
    import atexit
    from scraper import close_browser
    atexit.register(close_browser)

    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        scheduler = BackgroundScheduler(timezone='Asia/Seoul')
        scheduler.add_job(scheduled_job, 'cron', hour=8, minute=0, id='morning_update')
        scheduler.start()
        logger.info('스케줄러 시작: 매일 오전 8시 자동 조회')
    except ImportError:
        logger.warning('APScheduler 미설치')

    try:   # 시작 직후 시장금리 1회 갱신(키 있으면)
        threading.Thread(target=fetch_bond_rates, daemon=True).start()
    except Exception:
        logger.exception('시장금리 시작갱신 실패')

    from waitress import serve
    serve(app, host='0.0.0.0', port=PORT)
